import openpyxl, os 
from openpyxl import Workbook
import datetime
import time 
import getpass 

def count_time(start):
    time_sec = start
    time_hour, time_sec =  time_sec // 3600, time_sec % 3600
    time_min, time_sec = time_sec // 60, time_sec % 60
    time_hour = str(int(time_hour))
    time_min = str(int(time_min))
    time_sec = str(round(time_sec, 2))
    time_sec = time_sec.split('.')
    time_mil = time_sec[1]
    time_sec = time_sec[0]
    count_up_str = "{}:{}:{}.{}".format(time_hour.zfill(2),
                                                    time_min.zfill(2),
                                                    time_sec.zfill(2),
                                                    time_mil.zfill(2))
    return count_up_str

def secs2time(s):
    ms = int((s - int(s)) * 1000000)
    s = int(s)
    # Get rid of this line if s will never exceed 86400
    #while s >= 24*60*60:
        #s -= 24*60*60
    h = s / (60*60)
    s -= h*60*60
    m = s / 60
    s -= m*60
    return datetime.time(h, m, s, ms)

# function to get the required time for each distance
def time_calc(dist, tspeed):
    hour_in_seconds = 3600.0
    return round((dist / tspeed) * hour_in_seconds, 1)

def get_time_list(total, speed):
    current_dist = 0.0
    time_list = []

    #while (current_dist < total - .1):
    for i in range(0, total):
        current_dist += .1
        #time_list.append(str(secs2time(time_calc(current_dist, speed))))
        time_list.append(count_time(time_calc(current_dist, speed)))
    
    return time_list

def save_list_in_excel(pulse_list, target_list, mile_list, tire_size, target_miles, target_speed):
    # Create the spreadsheet 
    # filename = r"C:\Users\AIOTIK-005\Desktop\Route_created.xlsx"
    book = Workbook()
    ws = book.active

    # Reference
    ws['H1'] = "REF"

    # write the target data here
    ws['I1'] = target_speed
    ws['J1'] = "MPH"

    ws['K1'] = tire_size
    ws['L1'] = "Inch"
    
    ws['M1'] = target_miles
    ws['N1'] = "Miles"

    # setup colum A and B width
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    # set sheet name 
    ws.append(["PULSES", "TIME", "MILES"])

    # append the values to the column 
    # append the mile_list to the spreadsheet 
    r = 3
    for statN in mile_list:
        ws.cell(row = r, column = 3).value = statN
        r += 1
    
    # append the pulse list to the spreadsheet
    r = 3
    for statN in pulse_list:
        ws.cell(row = r, column = 1).value = '=(63360/K1/2) * {}'.format(str(r - 2))
        ws.cell(row = r, column = 1).number_format = '0'
        #ws.cell(row = r, column = 1).value = statN
        r += 1
    
    # append the target list to spreadsheet
    r = 3
    for statN in target_list:
        ws.cell(row = r, column = 2).value = '=((C{} / I1) * 3600.0) / 86400'.format(r)
        ws.cell(row = r, column = 2).number_format = 'hh:mm:ss.00'
        r += 1

    # create the spreadsheet 
    user = os.getcwd().split('\\')
    user = user[0] + '\\' + user[1] + '\\' + user[2]
    filename = str(input("Enter the name for the spreadsheet: "))
    filepath = user + "\Desktop\\" + filename + r'.xlsx'
    # filepath = r"C:\Users\\" + getpass.getuser() +  r"\Desktop\\" + filename + r".xlsx"
    print(f"filepath = {filepath}")
    book.save(filepath)

def calculate_pulses(tire_size):
    return (63360 / tire_size) / 2         # rotations per mile = (mile in inches / tire circumference) 
                                           # 6336 = 0.1 miles

def get_pulses_list(first_pulse, total):
    multiplier = 0
    pulse_list = []

    #while (multiplier < total):
    for i in range(0, total):
        #pulse_list.append(str(secs2time(time_calc(multiplier, speed))))
        multiplier += 1
        result = round(first_pulse * multiplier, 0)
        pulse_list.append(result)
    
    return pulse_list

def get_miles_list(total):
    mile_list = []
    mile_counter = 0

    while (mile_counter < total):
        mile_counter += .1
        mile_counter = round(mile_counter, 1)
        mile_list.append(mile_counter)
    
    return mile_list

# main application
def main(): 

    # a warning message 
    print("////////////////////////////////////////////////////")
    print("PLEASE, CLOSE ANY EXCEL FILE RELATED TO THIS PROGRAM")
    print("////////////////////////////////////////////////////\n")
    
    # in miles get the user data input 
    tire_size = float(input("Input the tire size (inches): "))
    total_length = float(input("Input total race length (miles): "))
    my_speed = float(input("Input target speed (mph): "))
    
    # get mile and time list generated 
    mile_list = get_miles_list(total_length)
    timer_list = get_time_list(len(mile_list), my_speed)
    
    # calculate pulses and get the pulses list
    pulses = calculate_pulses(tire_size)
    p_list = get_pulses_list(pulses, len(mile_list))
    
    # save the data into an spreadsheet
    save_list_in_excel(p_list, timer_list, mile_list, tire_size, total_length, my_speed)


# run the application
main()
print("\n////////////////////////////////////////////////////")
print("SPREADSHEET CREATED, NOW THIS WINDOW IS ABLE TO BE CLOSED")
print("////////////////////////////////////////////////////\n")
input("Enter any key to exit")
exit()
