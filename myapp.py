import kivy

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.lang import Builder

import os
import openpyxl

Builder.load_string("""
<MyWidget>:
    id: my_widget
    orientation: 'vertical'
    FileChooserListView:
        id: filechooser
        path: "/home/pi/Documents/myRoutes"
        on_selection: my_widget.selected(filechooser.selection)
    Button
        text: "open"
        on_release: my_widget.open(filechooser.path, filechooser.selection)
""")

class MyWidget(BoxLayout):
    def open(self, path, filename):
        with open(os.path.join(path, filename[0])) as f:
            # working very well!
            book = openpyxl.load_workbook(filename[0])
            sheet = book.active
            #cells = sheet['A1': 'B6']
            #cells = sheet[sheet.min_row : sheet.max_column]
            cells = sheet[sheet.dimensions]

            for c1, c2, c3 in sheet[sheet.dimensions]:
                #print(c1.value, c2.value)
                print("{} {} {}".format(c1.value, c2.value, c3.value))

            #for c1, c2 in cells:
            #    print("{0:8} {1:8}".format(c1.value, c2.value))

            # this is already working 
            #a1 = sheet['A1']
            #a2 = sheet['A2']
            #a3 = sheet.cell(row = 3, column = 1)
            #print(a1.value)
            #print(a2.value)
            #print(a3.value)
            

            # other stuff that is not important right now
            #book = f
            #print(filename)
            #book = openpyxl.load_workbook(f.read()
            #a = f.read() 
            #print a
            #print f.read()

    def selected(self, filename):
        print "selected: %s" % filename[0]


class MyApp(App):
    def build(self):
        return MyWidget()

if __name__ == '__main__':
    MyApp().run()