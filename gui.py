
# -*- coding: utf-8 -*-
# let's wait a bit before starting
import time

import Peripherals as peripherals
#import pulseCountLog as pulseFunc    # testing library
from kivy.app import App
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen, FadeTransition
from kivy.properties import ObjectProperty, StringProperty, NumericProperty
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import YELLOW
from openpyxl.styles.colors import BLACK
from kivy.properties import ListProperty
from kivy.uix.label import Label
from kivy.uix.widget import Widget
from kivy.graphics import Line
from kivy.uix.popup import Popup
from kivy.uix.gridlayout import GridLayout
from kivy.uix.textinput import TextInput
from kivy.clock import Clock
import os, sys, os.path, subprocess, glob

import openpyxl
from openpyxl import Workbook
import json
import re 
from geopy.distance import great_circle
import Peripherals as peripherals
import threading
from kivy.properties import BooleanProperty
from kivy.uix.togglebutton import ToggleButton
#from sympy import Integral, Symbol
import RPi.GPIO as IO
import math
import datetime
from kivy.clock import mainthread
from kivy.uix.button import Button



# Define GPIOs
RPM_SENSOR_PIN = 12#5
#TEMP_SENSOR_ADDR = 0x5A
# Define GPIO Mode
GPIO_MODE = IO.BCM


# Define other parameters
TIRE_RADIUS_REF = 0.0 #13.1 # Inch
TIRE_PRESSURE_REF = 0.0 #33 # Pressure of reference
TIRE_TEMPERATURE_REF = 0
TARGET_SPEED = 3
STOP_BUTTON = 1

sense = peripherals.Sensors
record = peripherals.Record
gps = peripherals.GPS
#dataToStrings = peripherals.dataToStrings
routes = peripherals.Routes
routes_data = routes()
SET_VALUE_RECORD = True
value_to_compare = 255

# bluetooth!
#ble = peripherals.Bluetoothctl

# Classes
# Sensors peripheral class
init_sense = sense(
GPIO_MODE,
RPM_SENSOR_PIN)#,
#TEMP_SENSOR_ADDR
#)
# Record class
init_record = record()
# Gps class
init_gps = gps()
# dataToStrings class
#init_strings = dataToStrings()
# Setup the GPS device
gps.setup_skytraq_gps(init_gps)
# Set temperature of reference
#sense.set_temperature_ref(init_sense)
#sense.temperature_ref["tire_F"] = Main_Screen.temp
#temp =
#pres =

#bluetooth!
#bl = ble()
#bl.start_scan()
#mac = "04:B3:EC:07:11:61"
#mac = "04:B3:EC:01:1E:65"
my_counter = 0
pres = "0"
temp = "0"


IO.output(record.green1_led, False)
IO.output(record.green2_led, False)
IO.output(record.green3_led, False)
IO.output(record.green4_led, False)
IO.output(record.green5_led, False)
IO.output(record.yellow_led, False)
IO.output(record.red1_led, False)
IO.output(record.red2_led, False)
IO.output(record.red3_led, False)
IO.output(record.red4_led, False)
IO.output(record.red5_led, False)
IO.output(record.gps_working, False)

# Create the GPS Thread to call GPS data ASAP
thread_flag = False    # to enable/disable the thread
start_system = 0

#clock_status = 0
main_loop = 0
input_loop = 0
static_message_loop = 0 
race_setup_screen_loop = 0

def function1():
    time.sleep(5)
    print("Done")
    Clock.schedule_once(App.get_running_app().show_main_screen)


def function2():
    time.sleep(5)
    print("Done 2")
    Clock.schedule_once(App.get_running_app().show_main_screen)

class Show_Val(Screen):
    message = StringProperty('')
    timer_update = StringProperty('')
    timer_2show_screen = StringProperty('')
    dist_travel = StringProperty('')
    count_down_up = StringProperty('') #average_mph = StringProperty('')
    temp_val = StringProperty('')
    dist_travel_2 = StringProperty('')
    tpms_temp = StringProperty('')
    tpms_pressure = StringProperty('')
    current_color = ListProperty()
    current_value = StringProperty('')
    show_me_data = StringProperty('')
    tire_size = StringProperty('')
    switch_ref = StringProperty('')
    gps_color = ListProperty()
    text_gps = StringProperty('')
    clear_text_input = StringProperty('')
    color_text_gps = ListProperty()
    press_indicator = StringProperty('')
    temperature_indicator = StringProperty('')
    enable_toggle_button = BooleanProperty(False)
    speed_target_indicator = StringProperty()
    pulseCounter = StringProperty('')
    saveRouteInUSB = StringProperty('')
    loadRouteInList = StringProperty('')
    rpi_temp = StringProperty('')
    hidePulseMeasure = BooleanProperty(False)

class File_List_To_Save_Usb(Screen):
    update_val = Show_Val()
    preloaded_target_speed = 0

    #waypoint = {'mph':0, 'distance measured': 0, 'latitude':0, 'longitude':0, 'hours':0, 'minutes':0, 'seconds':0}
    waypoint_mph_from_datalogger =       []
    waypoint_dist_meas_from_datalogger = []
    waypoint_latitude_from_datalogger =  []
    waypoint_longitude_from_datalogger = []
    waypoint_time                      = []
    waypoint_hours_from_datalogger =     []
    waypoint_minutes_from_datalogger =   []
    waypoint_seconds_from_datalogger =   []
    waypoint_counter_from_datalogger =   []
    waypoint_target_speed_from_datalogger = []
    old_target_speed = 0
    counter_from_datalogger_from_datalogger = 0
    last_dist_meas_from_datalogger = 0.0

    preloaded_route_done = 1
    uploaded = 0

    my_time_array = "0"
    my_time_complete = []

    excel_load_indicator = 0
    list_2 = []
    file_loaded = ""

    show_message = False
    user_popup = ""

    file_to_delete = ""
    file_to_load = ""
    file_already_loaded = False
    student_list = ObjectProperty()
    file_name = ""
    key_list_logger = False
    

    def __init__(self, **kwargs):
        super(File_List_To_Save_Usb, self).__init__(**kwargs)
        Clock.schedule_interval(self.update_list_data, 1)

    def update_list_data(self, dt):
        limitSavedRoutes = 10
        if File_List_To_Save_Usb.key_list_logger == True:
            list_1 = []
            index = -1
            out = subprocess.check_output("ls /home/pi/Documents/SaveRoute", shell=True)  # test for RPi
            sp = out.split(".xlsx")
            #print("sp {}".format(sp))
            if len(sp) >= 2:
                for i in range(len(sp)):
                    if(sp[i] != '\n'):
                        list_1.append(sp[i].strip())
                if len(list_1) > len(File_List_To_Save_Usb.list_2):
                    for i in list_1:
                        for x in File_List_To_Save_Usb.list_2:
                            if i == x: 
                                break
                        else:
                            self.student_list.adapter.data.extend([i])
                            File_List_To_Save_Usb.list_2.append(i)                
                if len(list_1) < len(File_List_To_Save_Usb.list_2):
                    print("current is smaller ")
                    for i in File_List_To_Save_Usb.list_2: # list 1
                        for x in list_1:    
                            if i == x: 
                                index += 1
                                break
                        else:
                            index += 1
                            print("item from list 1 {} is NOT in list 2 and I have to delete it and the index is {}".format(i,index))
                            selection = i
                            self.student_list.adapter.data.remove(selection)
                            del File_List_To_Save_Usb.list_2[index]
                if len(list_1) > limitSavedRoutes:
                    File_List_To_Save_Usb.update_val.saveRouteInUSB = "PLEASE, REMOVE A FEW ROUTES"
                else:
                    File_List_To_Save_Usb.update_val.saveRouteInUSB = "SELECT A ROUTE TO SAVE IN USB"    


    def check_name(self):
        if self.student_list.adapter.selection:
            File_List_To_Save_Usb.file_name = self.student_list.adapter.selection[0].text
            print("file to save {}".format(File_List_To_Save_Usb.file_name))
            guiApp.message_selected = Static_Message.NO_ERROR_IN_USB
        else:
            guiApp.message_selected = Static_Message.NO_ROUTE_SELECTED_IN_SAVE_MODE
    
    def uncheck(self):  # stop to look for a new drive and file routes!
        File_List_To_Save_Usb.key_list_logger = False
    
    def change_name(self, name_to_show):
        print("{}".format(len(name_to_show)))
        if name_to_show: # if file is selected
            if len(name_to_show) == 0:
                pass #p = CustomPopup_1()
                        #p.show_popup()
            else:
                name = name_to_show
                if len(name) >= 9:
                    print("little name {}...".format(name[:7]))
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name[:7] + "..."
                else:
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name
                print("name {}".format(name))

    def open_excel_file(self):
        if self.student_list.adapter.selection: # if file is selected
            Routes_List_Screen.file_to_load = self.student_list.adapter.selection[0].text
            print(Routes_List_Screen.file_to_load)
            #load_message = CustomPopup_5()
            #load_message.open()
        else:
            guiApp.message_selected = Static_Message.NO_SELECTED_ROUTE
            #p = CustomPopup_1()
            #p.show_popup()

class Input_Target_Speed_Screen(Screen):

    # target speed data
    target_speed = ObjectProperty()
    t_speed_variable = 0#StringProperty('')


    def speed_test(self):
        print("you made it")

    def t_speed(self):
        if(self.target_speed.text == "" or self.target_speed.text == "00" or self.target_speed.text == "000"):
            self.target_speed.text = "0"
            Input_Target_Speed_Screen.t_speed_variable = 0
        else:
            if "." in self.target_speed.text:
                text_filter = re.findall(r'^[a-zA-Z0-9,]*\.?[a-zA-Z0-9,]*$', self.target_speed.text)
                print("my text filter gets {}".format(text_filter))
                if len(text_filter) == 0 or text_filter[0] == ".":
                    print("you must show a popup")
                    self.target_speed.text = "0"
                    Input_Target_Speed_Screen.t_speed_variable = 0
                    guiApp.message_selected = Static_Message.INCORRECT_SPEED_FORMAT
                else:   
                    Input_Target_Speed_Screen.t_speed_variable = float(self.target_speed.text)
                    guiApp.message_selected = Static_Message.NO_ERROR
            else:
                record.general_purpose = 1
                Input_Target_Speed_Screen.t_speed_variable = float(self.target_speed.text)
                guiApp.message_selected = Static_Message.NO_ERROR
        print(Input_Target_Speed_Screen.t_speed_variable)


class Routes_List_Screen(Screen):
    update_val = Show_Val()
    preloaded_target_speed = 0
    waypoint_mph_from_datalogger =       []
    waypoint_dist_meas_from_datalogger = []
    waypoint_latitude_from_datalogger =  []
    waypoint_longitude_from_datalogger = []
    waypoint_time                      = []
    waypoint_hours_from_datalogger =     []
    waypoint_minutes_from_datalogger =   []
    waypoint_seconds_from_datalogger =   []
    waypoint_counter_from_datalogger =   []
    waypoint_target_speed_from_datalogger = []
    old_target_speed = 0
    counter_from_datalogger_from_datalogger = 0
    last_dist_meas_from_datalogger = 0.0

    preloaded_route_done = 1
    uploaded = 0

    my_time_array = "0"
    my_time_complete = []

    excel_load_indicator = 0
    list_2 = []
    file_loaded = ""

    show_message = False

    file_to_delete = ""
    file_to_load = ""
    file_already_loaded = False
    #first_name_text_input = ObjectProperty()
    #last_name_text_input = ObjectProperty()
    student_list = ObjectProperty()
    key_file_list = False

    #def refresh_list(self):
    #    self.list_item._trigger_reset_populate()
    def __init__(self, **kwargs):
        super(Routes_List_Screen, self).__init__(**kwargs)
        Clock.schedule_interval(self.update_list_data, 1)

    def check_for_usb(self):
        #subprocess.call("sudo mount /dev/sda1 /media/pi/ -o uid=pi,gid=pi")
        out = subprocess.check_output("df", shell = True)   # Save devices plugged to device into "out"
        if(len(out) >= 570):                                # Check if the string is enoug to know if a drive is connected 
            print("step 1")
            if Main_Screen.sc_usb == 0:
                print("step 2")
                splited = out.split(" ")                        # Split the string every space
                for i in range(0, len(splited)):                # iterate the list
                    if(len(splited[i]) == 24):                  # if the item of the list is same as target item "sda1" device
                        usb_dev = splited[i].split("/")         # split the string every "/"
                        for i in range(0, len(usb_dev)):        # iterate current list
                            if((usb_dev[i] == "sda1") or (usb_dev[i] == "sdb1")):           # if iterated item is sda1 or sdb1
                                print("step 3, found an sd file ")
                                check_files = subprocess.check_output("ls /media/pi", shell = True) # save the output of 'ls /media/pi'
                                check_files = check_files.split("\n")                               # removing the "\n"
                                os.chdir("/media/pi/" + check_files[0])                             # take directory as reference
                                for file in glob.glob("*.xlsx"):                                     # Iterate each .txt file recognized
                                    # copy each file in the correspondent directory 
                                    #final = subprocess.check_output(("sudo cp /media/pi/" + check_files[0] + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)                             
                                    subprocess.call(("sudo cp /media/pi/" + '"' + check_files[0] + '"' + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)    
                                    print("done!")
                                Main_Screen.sc_usb = 1
                                Main_Screen.get_time_after_copy = time.time()
                                print("for loop done!")
    
    def update_list_data(self, dt):
        limitLoadFiles = 10
        if Routes_List_Screen.key_file_list == True:
            #self.check_for_usb()
            print("testing usb file")
            list_1 = []
            index = -1
            out = subprocess.check_output("ls /home/pi/Documents/LoadRoute", shell=True)  # test for RPi
            sp = out.split(".xlsx")

            for i in range(len(sp)):
                if(sp[i] != '\n'):
                    list_1.append(sp[i].strip())
            if len(list_1) > len(Routes_List_Screen.list_2):
                for i in list_1:
                    for x in Routes_List_Screen.list_2:
                        if i == x:  
                            break
                    else:
                        self.student_list.adapter.data.extend([i])
                        Routes_List_Screen.list_2.append(i)

            if len(list_1) < len(Routes_List_Screen.list_2):
                print("current is smaller ")
                for i in Routes_List_Screen.list_2: 
                    for x in list_1: 
                        if i == x: 
                            index += 1
                            break
                    else:
                        index += 1
                        print("item from list 1 {} is NOT in list 2 and I have to delete it and the index is {}".format(i,index))
                        selection = i
                        self.student_list.adapter.data.remove(selection)
                        del Routes_List_Screen.list_2[index]
            if len(list_1) > limitLoadFiles:
                Routes_List_Screen.update_val.loadRouteInList = "PLEASE, REMOVE A FEW ROUTES"
            else:
                Routes_List_Screen.update_val.loadRouteInList = "SELECT A ROUTE TO LOAD"    

    def delete_file(self):
        # If a list item is selected
        if self.student_list.adapter.selection:
            print(self.student_list.adapter.selection[0].text)
            Routes_List_Screen.file_to_delete = self.student_list.adapter.selection[0].text
            guiApp.message_selected = Static_Message.NO_ERROR
        else:
            guiApp.message_selected = Static_Message.NO_SELECTED_ROUTE
    
    def uncheck(self):  # stop to look for a new drive and file routes!
        #Main_Screen.look_for_files = False
        Routes_List_Screen.key_file_list = False
    
    def change_name(self, name_to_show):
        print("{}".format(len(name_to_show)))
        if name_to_show: # if file is selected
            if len(name_to_show) == 0:
                pass
            else:
                name = name_to_show
                if len(name) >= 9:
                    print("little name {}...".format(name[:7]))
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name[:7] + "..."
                else:
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name
                print("name {}".format(name))

    def get_file_name_to_load(self):
        if self.student_list.adapter.selection:
            print("this is the file: {}".format(self.student_list.adapter.selection))
            guiApp.message_selected = Static_Message.NO_ERROR
            Routes_List_Screen.file_to_load = self.student_list.adapter.selection[0].text
        else:
            print("could not find a file to load :(")
            guiApp.message_selected = Static_Message.NO_FILE_SELECTED 

class Main_Screen(Screen):
    screen_ch = time.time()

    screen_1 = 0
    screen_2 = 0
    screen_3 = 0
    screen_4 = 0
    screen_5 = 0
    update_val = Show_Val()
    label_wid = ObjectProperty()
    enabled = BooleanProperty(True)
    record_time = 0
    final_time = 0
    toggle_speed = ObjectProperty()
    togle_speed_str = StringProperty() 
    key_stop = False
    timeFrontDiff = 0
    dif_sec = {'hours': 0, 'minutes': 0, 'seconds': 0}
    result = 0
    pres = "0"
    pres_psi = "0"
    temp = "0"
    test = 0
    the_high = 0
    the_low = 0
    mode_selector = 1
    clear_comparison = 0

    my_iterator = 0

    comparison = 0.0
    saver_data = "0"
    yellow  =    [0.95, 0.98, 0, 1]
    green   =    [0.015, 0.98, 0, 1]
    red     =    [1, 0.16, 0.16, 1]
    black   =    [0, 0, 0, 1]
    white   =    [1, 1, 1, 1]

    update_val.current_color = yellow
    stop_iteration = 0
    start_now = 0
    gps_time_high = 0
    gps_time_low = 0

    ref_speed = 0
    ref_pressure = 0
    ref_inch = 0
    ref_temp = 0

    flag_gps_high = 0
    json_pres = []
    json_temp = []
    json_inch = []

    counter_test = 0
    locked_high = 0
    ref_size = 0
    IO.output(record.gps_working, False)
    start_system = 0
    file_path_name = '/home/pi/Documents/tireData/data_input.json'
    data = json.load(open(file_path_name))
    data_temp = data["temperature"]
    data_pres = data["pressure"]
    data_size = data["size"]
    mode_selector = 1   #name can be only 10 Characters
    update_val.text_gps = "NO ROUTE \nSELECTED"
    update_val.color_text_gps = black
    update_val.gps_color = yellow
    Routes_List_Screen.preloaded_route_done = 1
    Routes_List_Screen.uploaded  = 0
    flag_race_mode = 0
    stable_gps = time.time()
    stable_gps_result = 0
    measured_time = 0
    sense.mph_reference = sense.mph
    time_sec = 0
    time_hour = 0
    time_min = 0
    time_mil = 0
    count_up_str = '0'

    #sense.gps_thread = threading.Thread(target=gps_thread)  # instance the thread
    #sense.gps_thread.start()   # call to start the thread

    sc_usb = 0
    get_time_after_copy = 0
    measure_time_usb_copy = 0
    listed = []
    look_for_files = False
    lock_toggle = False

    last_ = 0

    spIndex = 0
    #calculate_time = 0
    current_race_time = 0
    button = ObjectProperty(None)
    bbuton = 0
    toggle_changer_available = False
    bool_race_leg = False
    bool_reset_leg = False
    bool_load_time_leg = False
    leg_time = 0
    bool_leg_already_elapsed = False
    end1stLegTimeDiff = 0
    end1stLegTime = 0
    add_2nd_leg = False
    key_1st_leg_end = False
    rpi_string = "0"
    
    time_temp_last = 0
    time_temp_result = 0
    time_temp_current = time.time()

    def pressed_button(self):   # change the mph reference measurement to gps module
        print('SPEED REFERENCE: GPS')
        sense.switch_mph_reference = True
        sense.mph = 0

    def unpressed_button(self): # change the mph reference measurement to hall effect sensor
        print('SPEED REFERENCE: RPM')
        sense.switch_mph_reference = False
        sense.mph = 0

    def race_setup_mode(self):
        Main_Screen.mode_selector = 1
        Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
        Main_Screen.update_val.color_text_gps = Main_Screen.black
        Main_Screen.update_val.gps_color = Main_Screen.yellow
        Routes_List_Screen.preloaded_route_done = 1
        Routes_List_Screen.uploaded  = 0

    def check_state(self):
        if (sense.race_started == True):
            return False
        else:
            return True

    def race_mode(self):
        Main_Screen.mode_selector = 0
        
    def count_time(self, start):
        Main_Screen.time_sec = start
        Main_Screen.time_hour, Main_Screen.time_sec =  Main_Screen.time_sec // 3600, Main_Screen.time_sec % 3600
        Main_Screen.time_min, Main_Screen.time_sec = Main_Screen.time_sec // 60, Main_Screen.time_sec % 60
        Main_Screen.time_hour = str(int(Main_Screen.time_hour))
        Main_Screen.time_min = str(int(Main_Screen.time_min))
        Main_Screen.time_sec = str(round(Main_Screen.time_sec, 2))
        Main_Screen.time_sec = Main_Screen.time_sec.split('.')
        Main_Screen.time_mil = Main_Screen.time_sec[1]
        Main_Screen.time_sec = Main_Screen.time_sec[0]
        Main_Screen.count_up_str = "{}:{}:{}.{}".format( Main_Screen.time_hour.zfill(2),
                                                        Main_Screen.time_min.zfill(2),
                                                        Main_Screen.time_sec.zfill(2),
                                                        Main_Screen.time_mil.zfill(2))
        return Main_Screen.count_up_str

    def new_calculation_speed(self, pressure, temp):
        global TIRE_RADIUS_REF
        Main_Screen.ref_pressure = pressure
        Main_Screen.ref_temp = temp
        for current in range(0, len(Main_Screen.data_pres)):
            if(int(Main_Screen.ref_pressure) == int(Main_Screen.data_pres[current]) and int(Main_Screen.data_pres[current]) != 0):
                if(int(Main_Screen.ref_temp) == int(Main_Screen.data_temp[current]) and int(Main_Screen.data_temp[current]) != 0):
                    print("ref_temp: {} is equal to data_temp: {}".format(Main_Screen.ref_temp, int(Main_Screen.data_temp[current])))
                    Main_Screen.ref_size = float(Main_Screen.data_size[current])
                    print("data_size: {}".format(Main_Screen.data_size[current]))
                    TIRE_RADIUS_REF = Main_Screen.ref_size
                    #return Main_Screen.ref_size

    def check_for_usb(self):
        out = subprocess.check_output("df", shell = True)   # Save devices plugged to device into "out"
        if(len(out) >= 570):                                # Check if the string is enoug to know if a drive is connected 
            if Main_Screen.sc_usb == 0:
                splited = out.split(" ")                        # Split the string every space
                for i in range(0, len(splited)):                # iterate the list
                    if(len(splited[i]) == 24):                  # if the item of the list is same as target item "sda1" device
                        usb_dev = splited[i].split("/")         # split the string every "/"
                        for i in range(0, len(usb_dev)):        # iterate current list
                            if((usb_dev[i] == "sda1") or (usb_dev[i] == "sdb1")):           # if iterated item is sda1 or sdb1
                                check_files = subprocess.check_output("ls /media/pi", shell = True) # save the output of 'ls /media/pi'
                                check_files = check_files.split("\n")                               # removing the "\n"
                                os.chdir("/media/pi/" + check_files[0])                             # take directory as reference
                                for file in glob.glob("*.xlsx"):                                     # Iterate each .txt file recognized
                                    # copy each file in the correspondent directory 
                                    #final = subprocess.check_output(("sudo cp /media/pi/" + check_files[0] + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)                             
                                    subprocess.call(("sudo cp /media/pi/" + '"' + check_files[0] + '"' + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)                             
                                    print("done!")
                                Main_Screen.sc_usb = 1
                                Main_Screen.get_time_after_copy = time.time()
                                print("for loop done!")

    def clearRaceConditions(self):
        global STOP_BUTTON
        STOP_BUTTON = 1
        del routes.waypoint_latitude[:]
        del routes.waypoint_longitude[:]
        record.tm = 0
        record.record = False
        record.recording_status_flag = False
        record.toggle_status_flag = False
        record.t = 0
        record.recorded_time["seconds"] = 0.000
        record.recorded_time["minutes"] = 0
        record.recorded_time["hours"] = 0
        record.clear_timer = 0
        record.set_record_time = 0.001
        record.toggle_button_time = 0.0
        del routes.waypoint_hours[:]
        del routes.waypoint_minutes[:]
        del routes.waypoint_seconds[:]
        record.flag_go = 0
        sense.pulses = 0
        sense.mph = 0
        Main_Screen.spIndex = 0
        sense.factor = 1
        sense.decimal_factor = 1
        sense.lastPulse = 0
        Main_Screen.lock_toggle = False
        sense.last_ = 0
        Main_Screen.stable_gps = time.time()
        sense.race_started = False
        routes.counter = 0.0
        sense.dist_meas = 0.0
        routes.last_dist_meas = 0.0
        record.general_purpose = 0
        del routes.waypoint_mph[:]
        del routes.waypoint_dist_meas[:]
        del routes.waypoint_counter[:]
        del routes.waypoint_comparison_data[:]
        Main_Screen.update_val.current_value = "0"
        Main_Screen.update_val.current_color = Main_Screen.yellow
        Main_Screen.update_val.dist_travel_2 =  "0" + "mi"
        Main_Screen.update_val.dist_travel = "0" + " mi"
        #Main_Screen.update_val.count_down_up = "0" + " MPH"
        Main_Screen.update_val.count_down_up = "0" + " Secs"
        Main_Screen.update_val.timer_update = "0:0:0.0"
        Main_Screen.flag_race_mode = 0
        routes.last_dist_meas = 0
        Main_Screen.my_iterator = 0
        Routes_List_Screen.preloaded_route_done = 1 
        Routes_List_Screen.uploaded = 0
        Main_Screen.result = 0
        record.starter_switch = 0
        Main_Screen.toggle_changer_available  = True
        Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
        IO.output(record.green1_led, False)
        IO.output(record.green2_led, False)
        IO.output(record.green3_led, False)
        IO.output(record.green4_led, False)
        IO.output(record.green5_led, False)
        IO.output(record.yellow_led, False)
        IO.output(record.red1_led, False)
        IO.output(record.red2_led, False)
        IO.output(record.red3_led, False)
        IO.output(record.red4_led, False)
        IO.output(record.red5_led, False)
        IO.output(record.gps_working, False)

    def raceCompare(self):
        pass
    
    def restartLegBase(self):
        global STOP_BUTTON
        record.starter_switch = 0
        STOP_BUTTON = 1
        #Main_Screen.comparison = 0
        record.the_start = 0
        record.the_end = 0
        record.dont_change_switch_opt = False
        #sense.pulses = 0
        routes.counter = 0.0
        #sense.dist_meas = 0.0
        #routes.last_dist_meas = 0.0
        record.general_purpose = 0
        Main_Screen.start_now = 0
        #del routes.waypoint_mph[:]
        #del routes.waypoint_dist_meas[:]
        #del routes.waypoint_counter[:]
        Main_Screen.lock_toggle = False
        record.rst_time_lock = False
        record.tm = 0
        record.last_time_gps = 0
        record.reverse_count = 0
        record.record = False
        record.recording_status_flag = False
        record.toggle_status_flag = False
        record.t = 0
        record.recorded_time["seconds"] = 0.000
        record.recorded_time["minutes"] = 0
        record.recorded_time["hours"] = 0
        record.clear_timer = 0
        record.set_record_time = 0.001
        record.toggle_button_time = 0.0
        record.gps_reference = 0
        record.starter_switch = 0
        record.substracter_time = int(record.sw_time)
        record.tm = 0
        sense.race_started = False
        routes.last_dist_meas = 0 
        sense.flag_to_measure_dist = False 
        Main_Screen.flag_race_mode = 0
        #del routes.waypoint_hours[:]
        #del routes.waypoint_minutes[:]
        #del routes.waypoint_seconds[:]
        self.update_val.enable_toggle_button = False 
        record.starter_time = 0

    def compareRoutes(self):
        Main_Screen.dif_sec["seconds"] = Main_Screen.comparison
        # when t1 is greater than t2
        if(Main_Screen.dif_sec['seconds'] > 0):
            self.update_val.current_color = Main_Screen.red
            # when is greater than 60 seconds
            if(Main_Screen.dif_sec['seconds'] >= 60):
                Main_Screen.dif_sec['minutes'], Main_Screen.dif_sec['seconds'] = divmod(Main_Screen.dif_sec['seconds'], 60)
                Main_Screen.dif_sec['seconds'] = round(Main_Screen.dif_sec['seconds'], 3)
                Main_Screen.result = ("+{}:{}".format(int(Main_Screen.dif_sec['minutes']), round(Main_Screen.dif_sec['seconds'], 2)))
                # when is greater than 60 minutes
                if(int(Main_Screen.dif_sec['minutes']) > 59):
                    Main_Screen.dif_sec['hours'], Main_Screen.dif_sec['minutes'] = divmod(Main_Screen.dif_sec['minutes'], 60)
                    Main_Screen.result = ("+{}:{}:{}".format(int(Main_Screen.dif_sec['hours']), int(Main_Screen.dif_sec['minutes']), round(Main_Screen.dif_sec['seconds'], 2)))
            # when is only about seconds
            else:
                Main_Screen.result = '+' + str(Main_Screen.dif_sec['seconds'])
        # when t1 is equal to t2
        elif(Main_Screen.dif_sec["seconds"] == 0):
            self.update_val.current_color = Main_Screen.yellow
            Main_Screen.result = str(Main_Screen.dif_sec['seconds'])
        # when t1 is less than t2
        else:
            # when is greater than 60 seconds
            self.update_val.current_color = Main_Screen.green
            if(Main_Screen.dif_sec['seconds'] <= (-60)):
                Main_Screen.dif_sec['minutes'], Main_Screen.dif_sec['seconds'] = divmod(Main_Screen.dif_sec['seconds'], (-60))
                Main_Screen.dif_sec['seconds'] = round(Main_Screen.dif_sec['seconds'], 3)
                no_sign = str(Main_Screen.dif_sec['seconds']).split('-')
                Main_Screen.result = ("-{}:{}".format(int(Main_Screen.dif_sec['minutes']), no_sign[1]))
                # when is greater than 60 minutes
                if(int(Main_Screen.dif_sec['minutes']) > 59):
                    Main_Screen.dif_sec['hours'], Main_Screen.dif_sec['minutes'] = divmod(Main_Screen.dif_sec['minutes'], (60))
                    no_sign = str(Main_Screen.dif_sec['seconds']).split('-')
                    Main_Screen.result = ("-{}:{}:{}".format(int(Main_Screen.dif_sec['hours']), int(Main_Screen.dif_sec['minutes']), no_sign[1]))
                    print(Main_Screen.result)#
            # when is only about seconds
            else:
                Main_Screen.result = str(Main_Screen.dif_sec['seconds']).split('-')
                Main_Screen.result = '-' + Main_Screen.result[1]
        if Main_Screen.comparison >= .1:
            IO.output(record.red1_led, True)
            if Main_Screen.comparison >= .2: #or Main_Screen.dif_sec['minutes'] > 1 or Main_Screen.dif_sec['hours'] > 1):
                IO.output(record.red2_led, True)
                if Main_Screen.comparison >= .3: # or Main_Screen.dif_sec['minutes'] > 1 or Main_Screen.dif_sec['hours'] > 1):
                    IO.output(record.red3_led, True)
                    if Main_Screen.comparison >= .4: # or Main_Screen.dif_sec['minutes'] > 1 or Main_Screen.dif_sec['hours'] > 1):
                        IO.output(record.red4_led, True)
                        if Main_Screen.comparison >= .5: # or Main_Screen.dif_sec['minutes'] > 1 or Main_Screen.dif_sec['hours'] > 1):
                            IO.output(record.red5_led, True)
                        else:
                            IO.output(record.red5_led, False)
                    else:
                        IO.output(record.red4_led, False)
                        IO.output(record.red5_led, False)
                else:
                    IO.output(record.red3_led, False)
                    IO.output(record.red4_led, False)
                    IO.output(record.red5_led, False)
            else:
                IO.output(record.red2_led, False)
                IO.output(record.red3_led, False)
                IO.output(record.red4_led, False)
                IO.output(record.red5_led, False)
        else:
            IO.output(record.red1_led, False)
            IO.output(record.red2_led, False)
            IO.output(record.red3_led, False)
            IO.output(record.red4_led, False)
            IO.output(record.red5_led, False)
        if Main_Screen.comparison == 0:
            IO.output(record.yellow_led, True)
        else:
            IO.output(record.yellow_led, False)
        if Main_Screen.comparison <= (-.1):
            IO.output(record.green1_led, True)
            if Main_Screen.comparison <= (-.2):
                IO.output(record.green2_led, True)
                if Main_Screen.comparison <= (-.3):
                    IO.output(record.green3_led, True)
                    if Main_Screen.comparison <= (-.4):
                        IO.output(record.green4_led, True)
                        if Main_Screen.comparison <= (-.5):
                            IO.output(record.green5_led, True)
                        else:
                            IO.output(record.green5_led, False)
                    else:
                        IO.output(record.green4_led, False)
                        IO.output(record.green5_led, False)
                else:
                    IO.output(record.green4_led, False)
                    IO.output(record.green3_led, False)
                    IO.output(record.green5_led, False)
            else:
                IO.output(record.green4_led, False)
                IO.output(record.green3_led, False)
                IO.output(record.green2_led, False)
                IO.output(record.green5_led, False)
        else:
            IO.output(record.green4_led, False)
            IO.output(record.green3_led, False)
            IO.output(record.green2_led, False)
            IO.output(record.green1_led, False)
            IO.output(record.green5_led, False)

    def update(self, dt):
        global TIRE_RADIUS_REF
        global start_system
        global STOP_BUTTON
        global main_loop, input_loop, static_message_loop, race_setup_screen_loop
        try:
            # print("pulses: {}".format(sense.pulses))
            # ### TODO: IS REQUIRED TO ANALYZE WHY I did call this function 
            # POSSIBLY TO FIND A WAY TO CHANGE FROM SCREENS OVER PYTHON

            # if (time.time() - Main_Screen.screen_ch > 30):
            #     print("30 seconds elapased")
            #     Main_Screen.screen_ch = time.time()

            #     function2()
            
            #print("this is update and it's screen_ch {}".format(Main_Screen.screen_ch))
            Main_Screen.time_temp_current = time.time()
            Main_Screen.time_temp_result = int(Main_Screen.time_temp_current - Main_Screen.time_temp_last)
            #print("result {}".format(Main_Screen.time_temp_result))
            if Main_Screen.time_temp_result > 120:
                Main_Screen.time_temp_last = time.time()
                Main_Screen.rpi_string = subprocess.check_output("/opt/vc/bin/vcgencmd measure_temp", shell = True)
                Main_Screen.rpi_string = Main_Screen.rpi_string[5:]
                #print("string: {}".format(Main_Screen.rpi_string))
                self.update_val.rpi_temp = Main_Screen.rpi_string

            hours   = str(int(gps.hours))
            minutes = str(int(gps.minutes))
            seconds = str(gps.seconds)
            seconds = seconds.split('.')
            timer2disp = "{}:{}:{}".format(hours.zfill(2), minutes.zfill(2), seconds[0].zfill(2))
            #print("legs: {}".format(Confirmation_Screen_To_Lsense.end1stLegTimeDiff = insense.end1stLegTime)
            #print("end1stLegTimeDiff {}".format(Main_Screen.end1stLegTimeDiff))
            Main_Screen.end1stLegTimeDiff = int(time.time() - Main_Screen.end1stLegTime)
            if Main_Screen.bool_race_leg == True:
                print("evaluate") 
                #print("time time {}".format(Main_Screen.end1stLegTimeDiff))
                if Main_Screen.end1stLegTimeDiff < 15:
                    #print("asdassadawda")
                    #self.update_val.timer_update = "End 1st Leg"
                    self.update_val.dist_travel = str(sense.dist_meas) + " mi"
                    sense.mph = 0
                    self.update_val.count_down_up = "End 1st Leg" #str(sense.mph) + " MPH"
                    self.update_val.current_value = str(Main_Screen.result)
                    STOP_BUTTON = 0
                    sense.startReadPulses = False
                elif Main_Screen.end1stLegTimeDiff > 10:
                    Main_Screen.bool_reset_leg = True
                if Main_Screen.bool_reset_leg == True:
                    Main_Screen.bool_race_leg = False
                    Main_Screen.bool_reset_leg = False
                    Main_Screen.bool_load_time_leg = True
                    Main_Screen.leg_time = record.starter_time
                    self.restartLegBase()
            
                
            
            if Race_Setup_Screen.startRecordPulse == False: # if you are measuring pulses instead of race time
                if(Main_Screen.stable_gps_result >= 20):
                    self.update_val.timer_2show_screen = timer2disp
                else:
                    self.update_val.timer_2show_screen = "0:00:00"
                    self.update_val.timer_update = "0:00:00"
                    #self.update_val.timer_update = "Wait for GPS"
                    self.update_val.count_down_up = "Wait for GPS"

                if Main_Screen.key_stop == True:
                    set_correct_time = Main_Screen.current_race_time - Main_Screen.final_time
                    self.update_val.timer_update = self.count_time(Main_Screen.current_race_time - set_correct_time) 
                    #self.update_val.timer_update = "Race Ended"
                    self.update_val.count_down_up = "Race Ended"
                    Main_Screen.key_stop = False 
                #self.update_val.current_value = str(Main_Screen.comparison)
                self.update_val.current_value = str(Main_Screen.result)

                if record.starter_switch != 0:
                    if Main_Screen.lock_toggle == False:
                        if record.starter_time > 0:
                            if abs(record.starter_time) < 5:
                                self.update_val.enable_toggle_button = True 
                                Main_Screen.lock_toggle = True
                            else:
                                self.update_val.enable_toggle_button = False 

            Main_Screen.measure_time_usb_copy = time.time() - Main_Screen.get_time_after_copy
            if Main_Screen.measure_time_usb_copy >= 30:
                Main_Screen.sc_usb = 0
            
            if Main_Screen.toggle_changer_available == True:
                self.update_val.enable_toggle_button = False
                Main_Screen.toggle_changer_available = False

            
            # if gps is working!
            if(int(gps.latitude) != 0 and int(gps.longitude != 0)):
                Main_Screen.flag_gps_high = 1
                if(round((time.time() - Main_Screen.gps_time_high), 2) >= .5 and Main_Screen.the_high == 1):
                    IO.output(record.gps_working, True)
                    Main_Screen.gps_time_high = time.time()
                    Main_Screen.gps_time_low = time.time()
                    Main_Screen.the_high = 0
               
                if(round((time.time() - Main_Screen.gps_time_low), 2) >= .5 and Main_Screen.the_high == 0):
                    #print("blink OFF!")
                    IO.output(record.gps_working, False)
                    Main_Screen.gps_time_high = time.time()
                    Main_Screen.the_high = 1
            else:   # if there is no gps data valid, then GPS is not working
                Main_Screen.flag_gps_high = 0
                Main_Screen.locked_high == 0
                #print("latitude and longitude are not valid values")
                IO.output(record.gps_working, False)
            if (STOP_BUTTON == 1):
                # testing the speed measurement
                self.update_val.switch_ref = str(record.current_switch_time)
                record.set_record(init_record, SET_VALUE_RECORD) # Call this function to set Record button status
                record.start_record(init_record) # Call this function to start recording if all condition succeeds
                record.timer_switch(init_record)
                
                self.update_val.dist_travel = str(sense.dist_meas) + " mi"
                #self.update_val.count_down_up = str(sense.mph) + " MPH" #str(sense.mph_reference) + " MPH" #str(sense.mph) + " MPH"
                clear_time_hold = record.the_end - record.the_start
                if(clear_time_hold >= 5):
                    Main_Screen.comparison = 0
                    clear_time_hold = 0
                    record.the_start = 0
                    record.the_end = 0
                    record.dont_change_switch_opt = False
                    sense.pulses = 0
                    routes.counter = 0.0
                    sense.dist_meas = 0.0
                    routes.last_dist_meas = 0.0
                    record.general_purpose = 0
                    Main_Screen.start_now = 0
                    #del routes.waypoint_mph[:]
                    del routes.waypoint_dist_meas[:]
                    del routes.waypoint_counter[:]
                    del routes.waypoint_comparison_data[:]
                    Main_Screen.lock_toggle = False
                    record.rst_time_lock = False
                    record.tm = 0
                    record.last_time_gps = 0
                    record.reverse_count = 0
                    record.record = False
                    record.recording_status_flag = False
                    record.toggle_status_flag = False
                    record.t = 0
                    record.recorded_time["seconds"] = 0.000
                    record.recorded_time["minutes"] = 0
                    record.recorded_time["hours"] = 0
                    record.clear_timer = 0
                    record.set_record_time = 0.001
                    record.toggle_button_time = 0.0
                    record.gps_reference = 0
                    record.starter_switch = 0
                    record.substracter_time = int(record.sw_time)
                    record.tm = 0
                    sense.race_started = False
                    routes.last_dist_meas = 0 
                    sense.flag_to_measure_dist = False 
                    Main_Screen.flag_race_mode = 0
                    del routes.waypoint_hours[:]
                    del routes.waypoint_minutes[:]
                    del routes.waypoint_seconds[:]
                    self.update_val.enable_toggle_button = False 
                    record.starter_time = 0
                #bluetooth!

                #bl.data_ble = bl.get_device_info(mac)
                #bl.ble_counter += 1

                #if(bl.ble_counter >= 3):
                #    for i in range(0, len(bl.data_ble)):
                #        #if(len(data[i]) == 67):
                #        if(len(bl.data_ble[i]) == 67):
                #            Main_Screen.saver_data = bl.data_ble[i] # for reverse-engineering
                #            finder = bl.data_ble[i].split(" ")
                #            # testing
                #            #Main_Screen.pres = int(("0x" + finder[3]), 16) + int(("0x" + finder[4]), 16) + int(("0x" + finder[5]), 16) + int(("0x" + finder[6]), 16)
                #            #Main_Screen.pres_psi = round((float(Main_Screen.pres) / 6.89475729), 2)
                #            #Main_Screen.temp = int(("0x" + finder[7]), 16) + int(("0x" + finder[8]), 16)

                #if(Main_Screen.mode_selector == 1):
                record.time_string = str(gps.hours) + ":" + str(gps.minutes) + ":" + str(gps.seconds)
                record.time_string = format(record.time_string, '0<11')
                seconds_str = str(int(gps.seconds))
                seconds_str.zfill(2)

                if Race_Setup_Screen.startRecordPulse == False:
                    # print("1")
                    if(record.starter_switch == 0):
                        record.substracter_time = int(record.sw_time)
                        if(Main_Screen.stable_gps_result >= 20):
                            #self.update_val.timer_update = seconds_str.zfill(2) + " Secs"
                            self.update_val.count_down_up = seconds_str.zfill(2) + " Secs"
                            

                    if(record.starter_switch == 1 or record.starter_switch == 2):
                        record.starter_switch = 3
                        record.gps_reference = gps.seconds
                        #self.update_val.timer_update = "-" + str(record.starter_time) + " Secs"
                        self.update_val.count_down_up = "-" + str(record.starter_time) + " Secs"

                    
                    if(record.starter_switch != 0):
                        record.dont_change_switch_opt = True    # You cant change the switch selector on this point! 
                        if(record.sw_time == 60 or record.sw_time == 30):
                            if(record.gps_reference >= 30 and record.sw_time == 30):
                                r2 = record.gps_reference - 30
                                record.starter_time = record.sw_time - r2 - record.tm
                                record.substracter_time = record.sw_time - r2
                                if(record.starter_time > 0):
                                    if(int(record.starter_time) == 0):
                                        #self.update_val.timer_update = "0 " + "Secs"
                                        self.update_val.count_down_up = "0 " + "Secs"
                                    else:
                                        #self.update_val.timer_update = "-" + str(int(record.starter_time)).zfill(2) + " Secs"
                                        self.update_val.count_down_up = "-" + str(int(record.starter_time)).zfill(2) + " Secs"
                                else:
                                    record.rst_time_lock = True
                                    sense.race_started = True
                                    record.starter_time = -record.starter_time
                                    record.starter_time = record.starter_time + Main_Screen.leg_time 
                                    # if Main_Screen.leg_time > 0:
                                    #    record.starter_time = record.starter_time + Main_Screen.leg_time
                                    #    Main_Screen.leg_time = 0

                                    record.starter_time = round(record.starter_time, 2)
                                    sense.from_seconds_to_hours_gps = round((record.starter_time * 3600), 2)
                                    #if record.numLeg == 2 and Main_Screen.add_2nd_leg == True:
                                    #    record.starter_time = Main_Screen.leg_time
                                    #    Main_Screen.add_2nd_leg = False
                                    Main_Screen.current_race_time = record.starter_time + .99999
                                    self.update_val.timer_update = self.count_time(Main_Screen.current_race_time)
                            else:
                                record.starter_time = record.sw_time - record.gps_reference - record.tm
                                record.substracter_time = record.sw_time - record.gps_reference

                                if(record.starter_time > 0):
                                    if(int(record.starter_time) == 0):
                                        #self.update_val.timer_update = "0 " + "Secs"#str(int(record.starter_time)).zfill(2) + " Secs"
                                        self.update_val.count_down_up = "0 " + "Secs"#str(int(record.starter_time)).zfill(2) + " Secs"
                                    else:
                                        #self.update_val.timer_update = "-" + str(int(record.starter_time)).zfill(2) + " Secs"
                                        self.update_val.count_down_up = "-" + str(int(record.starter_time)).zfill(2) + " Secs"
                                else:
                                    record.rst_time_lock = True
                                    sense.race_started = True
                                    record.starter_time = -record.starter_time
                                    record.starter_time = record.starter_time + Main_Screen.leg_time 
                                    #if Main_Screen.leg_time > 0:
                                    #    record.starter_time = record.starter_time + Main_Screen.leg_time
                                    #    Main_Screen.leg_time = 0
                                    #record.starter_time = record.starter_time + Main_Screen.leg_time
                                    record.starter_time = round(record.starter_time, 2)
                                    sense.from_seconds_to_hours_gps = round((record.starter_time * 3600), 2)
                                    sense.flag_to_measure_dist = True 
                                    #if record.numLeg == 2 and Main_Screen.add_2nd_leg == True:
                                    #    record.starter_time = Main_Screen.leg_time
                                    #    Main_Screen.add_2nd_leg = False
                                    Main_Screen.current_race_time = record.starter_time + .99999
                                    self.update_val.timer_update = self.count_time(Main_Screen.current_race_time)
                                    
                #print("result {} and pulses {}".format(Main_Screen.result, sense.pulses))
                #if record.numLeg == 1 or record.numLeg == None:
                #self.raceCompare()             
                #print("one leg and value: {}".format(record.numLeg))
                #print("type: {}".format(type(record.numLeg)))
                if(Main_Screen.flag_race_mode == 1):
                    print("flace race mode!!")
                    #print("waypoint dist meas: {} and waypoint mph: {}".format(round(routes.waypoint_dist_meas[-1], 2), round(routes.waypoint_mph[-1], 2)))
                    #if round(routes.waypoint_dist_meas[-1], 2) != 0 and round(routes.waypoint_mph[-1], 2) != 0:
                    #    Main_Screen.calculate_time = (round(routes.waypoint_dist_meas[-1], 2) / round(routes.waypoint_mph[-1], 2))
                    #    Main_Screen.calculate_time = Main_Screen.calculate_time * 3600 
                        #print("calculated time: {}".format(Main_Screen.calculate_time))
                    #print("list {} len {}".format(Routes_List_Screen.waypoint_dist_meas_from_datalogger, len(Routes_List_Screen.waypoint_dist_meas_from_datalogger)))
                    #print("iterator {} and num of lists {} result {} and pulses {}".format(Main_Screen.my_iterator, (len(Routes_List_Screen.waypoint_dist_meas_from_datalogger) - 1), Main_Screen.result, sense.pulses))
                    #if Main_Screen.my_iterator >= (len(Routes_List_Screen.waypoint_dist_meas_from_datalogger) - 1):
                    if Main_Screen.spIndex >= (len(Routes_List_Screen.waypoint_dist_meas_from_datalogger) - 1):
                        #global STOP_BUTTON
                        global STOP_BUTTON
                        STOP_BUTTON = 0
                        Main_Screen.key_stop = True
                        self.update_val.dist_travel = str(sense.dist_meas)
                        self.update_val.current_value = str(Main_Screen.comparison)
                        print("time!!: {}".format(str(Main_Screen.comparison)))
                        self.compareRoutes()
                        #routes.waypoint_comparison_data.append(Main_Screen.comparison)
                        routes.waypoint_comparison_data.append(str(Main_Screen.result))
                        self.update_val.count_down_up = "Race Ended"
                        self.update_val.count_down_up = "Race Ended"
                        #print("current data to input in spreadsheet about comparison {}".format(routes.waypoint_comparison_data))
                        #self.update_val.current_value = str(Main_Screen.result)
                        #print("last waypoint miles measured: {} current distance: {}".format(Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.my_iterator], sense.dist_meas))
                        print("FINISHED! Iterator {} is equal to last waypoint_counter: {} AND DISTANCE {}".format(Main_Screen.my_iterator, Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.my_iterator], sense.dist_meas))
                        main_loop.cancel()
                        input_loop.cancel()
                        #static_message_loop.cancel() 
                        race_setup_screen_loop.cancel()
                        #Main_Screen.screen_1.cancel()
                        #Main_Screen.screen_2.cancel()
                        #Main_Screen.screen_3.cancel()
                        #Main_Screen.screen_4.cancel()

                        
                    else:
                            
                        print("doing my stuff")
                        if (Routes_List_Screen.waypoint_time[Main_Screen.my_iterator] == None) or (Routes_List_Screen.waypoint_time[Main_Screen.my_iterator] == " "):
                            print("Routes_List_Screen.waypoint_time[Main_Screen.my_iterator] {}".format(Routes_List_Screen.waypoint_time[Main_Screen.my_iterator]))
                        else:
                            #Main_Screen.comparison = record.get_sec(init_record, str(Routes_List_Screen.waypoint_time[Main_Screen.my_iterator]))  - round(Main_Screen.calculate_time, 2)       
                            self.compareRoutes()
                            #routes.waypoint_comparison_data.append(Main_Screen.comparison)
                            routes.waypoint_comparison_data.append(str(Main_Screen.result))
                            print("current data to input in spreadsheet about comparison {}".format(routes.waypoint_comparison_data))
                        Main_Screen.my_iterator += 1
                        Main_Screen.flag_race_mode = 0
                        Main_Screen.key_1st_leg_end = True

                
                    self.update_val.current_value = str(Main_Screen.result)

                    #print("comparison result {}".format(Main_Screen.result))
                    #print("quantity of pulses: {}, speed {}, and speed comparison {} and miles {}".format(sense.pulseResult, sense.mph, self.update_val.current_value, sense.dist_meas))
            #    #Main_Screen.result = Main_Screen.comparison
            #    self.update_val.current_value = str(Main_Screen.result)
        except KeyboardInterrupt:
            global thread_flag
            thread_flag = True
            IO.cleanup()
            sys.exit
        except IOError:
            pass


class Route_Screen(Screen):

    def check_for_usb(self):
        out = subprocess.check_output("df", shell = True)   # Save devices plugged to device into "out"
        if(len(out) >= 570):                                # Check if the string is enoug to know if a drive is connected 
            if Main_Screen.sc_usb == 0:
                splited = out.split(" ")                        # Split the string every space
                for i in range(0, len(splited)):                # iterate the list
                    if(len(splited[i]) == 24):                  # if the item of the list is same as target item "sda1" device
                        usb_dev = splited[i].split("/")         # split the string every "/"
                        for i in range(0, len(usb_dev)):        # iterate current list
                            if((usb_dev[i] == "sda1") or (usb_dev[i] == "sdb1")):           # if iterated item is sda1 or sdb1
                                check_files = subprocess.check_output("ls /media/pi", shell = True) # save the output of 'ls /media/pi'
                                check_files = check_files.split("\n")                               # removing the "\n"
                                os.chdir("/media/pi/" + check_files[0])                             # take directory as reference
                                for file in glob.glob("*.xlsx"):                                     # Iterate each .txt file recognized
                                    print "copying from folder"
                                    # copy each file in the correspondent directory 
                                    #final = subprocess.check_output(("sudo cp /media/pi/" + check_files[0] + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)                             
                                    subprocess.call(("sudo cp /media/pi/" + '"' + check_files[0] + '"' + "/" + '"' + file + '"' + "  /home/pi/Documents/LoadRoute"), shell = True)                             
                                    print("done!")
                                Main_Screen.sc_usb = 1
                                Main_Screen.get_time_after_copy = time.time()
                                print("for loop done!")

    def check(self):    # start to look for new a new drive 
        Main_Screen.look_for_files = True

    def check_files_to_load(self):  # check for files inside the usb drive and the load folder
        Routes_List_Screen.key_file_list = True

    def check_save_routes(self):
        File_List_To_Save_Usb.key_list_logger = True
    
    
#    def stop_record(self):
#        global STOP_BUTTON
#        STOP_BUTTON = 0
#
#    def start_record(self):
#        p = pulseFunc.CountLogger(sense.pulses)
#        print("this is p:  {}".format(p.myPulse))
#        #global STOP_BUTTON
#        #STOP_BUTTON = 1
#        #clear_time_hold = 0
#        #record.the_start = 0
        #record.the_end = 0
        #record.dont_change_switch_opt = False
        #sense.pulses = 0
        #routes.counter = 0.0
        #sense.dist_meas = 0.0
        #routes.last_dist_meas = 0.0
        #record.general_purpose = 0
        #Main_Screen.start_now = 0
        #del routes.waypoint_mph[:]
        #del routes.waypoint_dist_meas[:]
        #del routes.waypoint_counter[:]
        #record.rst_time_lock = False
        #record.tm = 0
        #record.last_time_gps = 0
        #record.reverse_count = 0
        #record.record = False
        #record.recording_status_flag = False
        #record.toggle_status_flag = False
        #record.t = 0
        #record.recorded_time["seconds"] = 0.000
        #record.recorded_time["minutes"] = 0
        #record.recorded_time["hours"] = 0
        #record.clear_timer = 0
        #record.set_record_time = 0.001
        #record.toggle_button_time = 0.0
        #record.gps_reference = 0
        #record.starter_switch = 0
        #record.substracter_time = int(record.sw_time)
        #record.tm = 0
        #routes.last_dist_meas = 0 
        #sense.flag_to_measure_dist = False 
        #Main_Screen.flag_race_mode = 0
        #del routes.waypoint_hours[:]
        #del routes.waypoint_minutes[:]
        #del routes.waypoint_seconds[:]

class Confirmation_Screen_to_Save(Screen):
            
    def save_file(self):
        # let's show wait message till the file is saved 
        Save_File_Text_Input_Screen.message_manager = Static_Message.SAVING_LABEL
        guiApp.message_selected = Static_Message.SAVING_LABEL
        time.sleep(2)

        def save_excel():
            global STOP_BUTTON
            global main_loop, input_loop, static_message_loop, race_setup_screen_loop

            Save_File_Text_Input_Screen.message_manager = Static_Message.SAVING_LABEL
            guiApp.message_selected = Static_Message.SAVING_LABEL
            time.sleep(2)
            key = False
            book = Workbook(write_only=True)
            sheet1 = book.create_sheet() #book.active
            sheet1.title = "RACE SHEET"
            #sheet2 = book.create_sheet(title = "PRE-RECORDED ROUTE")
            ws = book.get_active_sheet()
            #print("this is the text file: {}".format(Save_File_Text_Input_Screen.file_to_save))
            #print("and this is the self.text: {}".format(len(Save_File_Text_Input_Screen.file_to_save)))
        
            if len(Save_File_Text_Input_Screen.file_to_save) == 0 or Save_File_Text_Input_Screen.file_to_save.isspace() == True:
                #print("there is nothing here")
                guiApp.message_selected = Static_Message.INVALID_NAME
            else:
                #print("there is something")
                #Routes_List_Screen.waypoint_target_speed = Input_Target_Speed_Screen.t_speed_variable

                #redFill = PatternFill(start_color='FFFF0000',
                #           end_color='FFFF0000',
                #           fill_type='solid')

                #blueFill = PatternFill(start_color = '0b3ef4',
                #            end_color = '0b3ef4',
                #            fill_type = 'solid')

                # Set fixed cell size to 17 from A-H
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 22
                ws.column_dimensions['C'].width = 22
                ws.column_dimensions['D'].width = 22
                ws.column_dimensions['E'].width = 22
                ws.column_dimensions['F'].width = 22
                ws.column_dimensions['G'].width = 22
                ws.column_dimensions['H'].width = 22
                #ws.column_dimensions['I'].width = 22
                #ws.column_dimensions['J'].width = 22
                #ws.column_dimensions['K'].width = 22
                #ws.column_dimensions['L'].width = 22
                time.sleep(1)
                #ws.row_dimensions[1].height = 70
                ws.append([#"WAYPOINT NUMBER", 
                           "PULSES",
                           "PRE-RECORDED PULSES",
                           #"MPH AVERAGE",
                           "DISTANCE MEASURED",
                           "LATITUDE",
                           "LONGITUDE",
                           "TIME",
                           "PRE-LOADED TIME",
                           "COMPARISON RESULTS"])
                           #"TARGET SPEED"])
                time.sleep(1)
                #sheet1.cell(row = 1, column = 1).value = 'WAYPOINT NUMBER'
                #sheet1.cell(row = 1, column = 1).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 1).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 2).value = 'PULSES'
                #sheet1.cell(row = 1, column = 2).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 2).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet2.row_dimensions[1].height = 70
                #sheet1.cell(row = 1, column = 3).value = 'PRE-RECORDED PULSES'
                #sheet1.cell(row = 1, column = 3).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 3).font = openpyxl.styles.Font(bold = True, italic = True, color = BLACK)
                #sheet1.column_dimensions['A'].width = 17

                #sheet1.cell(row = 1, column = 4).value = 'MPH AVERAGE'
                #sheet1.cell(row = 1, column = 4).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 4).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 5).value = 'DISTANCE MEASURED'
                #sheet1.cell(row = 1, column = 5).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 5).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 6).value = 'LATITUDE'
                #sheet1.cell(row = 1, column = 6).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 6).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 7).value = 'LONGITUDE'
                #sheet1.cell(row = 1, column = 7).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 7).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 8).value = 'TIME'
                #sheet1.cell(row = 1, column = 8).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 8).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)

                #sheet1.cell(row = 1, column = 9).value = 'PRE-LOADED TIME'
                #sheet1.cell(row = 1, column = 9).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 9).font = openpyxl.styles.Font(bold = True, italic = True, color = BLACK)
                #sheet2.column_dimensions['B'].width = 17

                #sheet1.cell(row = 1, column = 10).value = 'COMPARISON RESULTS'
                #sheet1.cell(row = 1, column = 10).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 10).font = openpyxl.styles.Font(bold = True, italic = True, color = BLACK)

                #sheet1.cell(row = 1, column = 11).value = 'TARGET SPEED'
                #sheet1.cell(row = 1, column = 11).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet1.cell(row = 1, column = 11).font = openpyxl.styles.Font(bold = True, italic = False, color = BLACK)
    

                # Setting some style to each PRE RECORDED ROUTE
                #sheet2.row_dimensions[1].height = 70
                #sheet2.cell(row = 1, column = 1).value = 'PRE-RECORDED PULSES'
                #sheet2.cell(row = 1, column = 1).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet2.cell(row = 1, column = 1).font = openpyxl.styles.Font(bold = True, italic = True, color = BLACK)
                #sheet2.column_dimensions['A'].width = 17

                #sheet2.cell(row = 1, column = 2).value = 'PRE-RECORDED HH:MM:SS.ss'
                #sheet2.cell(row = 1, column = 2).alignment = openpyxl.styles.Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
                #sheet2.cell(row = 1, column = 2).font = openpyxl.styles.Font(bold = True, italic = True, color = BLACK)
                #sheet2.column_dimensions['B'].width = 17

                for i, x in enumerate(Routes_List_Screen.waypoint_time):
                    Routes_List_Screen.waypoint_time[i] = Routes_List_Screen.waypoint_time[i].strftime('%H:%M:%S.%f')

                for row in zip(
                    #routes.waypoint_counter,
                    routes.waypoint_pulses,
                    Routes_List_Screen.waypoint_dist_meas_from_datalogger,
                    #routes.waypoint_mph,
                    routes.waypoint_dist_meas,
                    routes.waypoint_latitude,
                    routes.waypoint_longitude,
                    routes.waypoint_seconds, 
                    Routes_List_Screen.waypoint_time,
                    routes.waypoint_comparison_data
                    ):
                    sheet1.append(row)
                time.sleep(2)
                #for iterator, distance_measured in enumerate(Routes_List_Screen.waypoint_dist_meas_from_datalogger):
                #    sheet2.cell(row = iterator+2, column = 1).value = distance_measured
                #for iterator, hours in enumerate(Routes_List_Screen.waypoint_time):
                #    sheet2.cell(row = iterator+2, column = 2).value = hours


                book.save('/home/pi/Documents/SaveRoute/' + Save_File_Text_Input_Screen.file_to_save + '.xlsx')
                

                #out = subprocess.check_output("df", shell = True)   # Save devices plugged to device into "out"
                #if(len(out) >= 570):                                # Check if the string is enoug to know if a drive is connected 
                #    #if sc_usb == 0:
                #    print("looking for a usb drive")
                #    splited = out.split(" ")                        # Split the string every space
                #    for i in range(0, len(splited)):                # iterate the list
                #        if(len(splited[i]) == 24):                  # if the item of the list is same as target item "sda1" device
                #            usb_dev = splited[i].split("/")         # split the string every "/"
                #            for i in range(0, len(usb_dev)):        # iterate current list
                #                if((usb_dev[i] == "sda1") or (usb_dev[i] == "sdb1")):           # if iterated item is sda1 or sdb1
                #                    print("{}".format(usb_dev[i]))
                #                    check_files = subprocess.check_output("ls /media/pi", shell = True) # save the output of 'ls /media/pi'
                #                    check_files = check_files.split("\n")                               # removing the "\n"
                #                    folder = subprocess.check_output("ls /media/pi/" + '"' +check_files[0] + '"', shell = True) # save the output of 'ls /media/pi'
                #                    folder = folder.split("\n")   
                #                    for i in range(len(folder)):
                #                        if folder[i] == "GPS SYSTEM GENERATED ROUTE":
                #                            print("I found the file {}".format(folder[i]))
                #                            key = True
                #                            break
#
                #                    if key == False:
                #                        folder_create = 'mkdir /media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'"
                #                        print("folder doesn't exist, I will create it")
                #                        print("{}".format(folder_create))
                #                        subprocess.call(folder_create, shell = True)
                #                        route = "/media/pi/" + '"' + check_files[0] + '"' + "/"
                #                        path = "cp /home/pi/Documents/SaveRoute/" + "'" + Save_File_Text_Input_Screen.file_to_save + '.xlsx' + "'"
                #                        destiny_folder = '/media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'" + "/"
                #                        print("folder to save 1: {}".format(path + " " + destiny_folder))
                #                        subprocess.call(path + " " + destiny_folder, shell = True)
                #                        Save_File_Text_Input_Screen.message_manager = Static_Message.SAVED_IN_SYSTEM_AND_USB
                #                    else:
                #                        key = False
                #                        path = "cp /home/pi/Documents/SaveRoute/" + "'" + Save_File_Text_Input_Screen.file_to_save + '.xlsx' + "'"
                #                        destiny_folder = '/media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'" + "/"
                #                        print("folder to save 2: {}".format(path + " " + destiny_folder))
                #                        subprocess.call(path + " " + destiny_folder, shell = True)
                #                        guiApp.message_selected = Static_Message.SAVED_IN_SYSTEM_AND_USB
                #                else:
                #                    guiApp.message_selected = Static_Message.SAVED_IN_SYSTEM
                #else:
                #    Save_File_Text_Input_Screen.message_manager = Static_Message.SAVED_IN_SYSTEM
                #    guiApp.message_selected = Static_Message.SAVED_IN_SYSTEM
                Save_File_Text_Input_Screen.message_manager = Static_Message.SAVED_IN_SYSTEM
                guiApp.message_selected = Static_Message.SAVED_IN_SYSTEM

                #print("value of message selected: {}".format(guiApp.message_selected))
                STOP_BUTTON = 1
                del routes.waypoint_latitude[:]
                del routes.waypoint_longitude[:]
                record.tm = 0
                Main_Screen.comparison = 0
                Main_Screen.leg_time = 0
                record.record = False
                record.recording_status_flag = False
                record.toggle_status_flag = False
                record.t = 0
                record.recorded_time["seconds"] = 0.000
                record.recorded_time["minutes"] = 0
                record.recorded_time["hours"] = 0
                record.clear_timer = 0
                record.set_record_time = 0.001
                record.toggle_button_time = 0.0
                del routes.waypoint_hours[:]
                del routes.waypoint_minutes[:]
                del routes.waypoint_seconds[:]
                record.flag_go = 0
                sense.pulses = 0
                sense.mph = 0
                sense.startReadPulses = False
                sense.counter_rotations = 0
                Main_Screen.spIndex = 0
                Main_Screen.bool_leg_already_elapsed = False
                sense.factor = 1
                sense.decimal_factor = round((sense.tire_rotations_per_mile * sense.factor), 0)
                sense.lastPulse = 0
                Main_Screen.lock_toggle = False
                sense.last_ = 0
                Main_Screen.stable_gps = time.time()
                sense.race_started = False
                routes.counter = 0.0
                sense.dist_meas = 0.0
                routes.last_dist_meas = 0.0
                record.general_purpose = 0
                #del routes.waypoint_mph[:]
                del routes.waypoint_dist_meas[:]
                del routes.waypoint_counter[:]
                del routes.waypoint_comparison_data[:]
                Main_Screen.update_val.current_value = "0"
                Main_Screen.update_val.current_color = Main_Screen.yellow
                Main_Screen.update_val.dist_travel_2 =  "0" + "mi"
                Main_Screen.update_val.dist_travel = "0" + " mi"
                #Main_Screen.update_val.count_down_up = "0" + " MPH"
                Main_Screen.update_val.count_down_up = "0" + " Secs"
                Main_Screen.update_val.timer_update = "0:0:0.0"
                Main_Screen.flag_race_mode = 0
                routes.last_dist_meas = 0
                Main_Screen.my_iterator = 0
                Routes_List_Screen.preloaded_route_done = 1 
                Routes_List_Screen.uploaded = 0
                Main_Screen.result = 0
                record.starter_switch = 0
                record.disablePulseMeas = False
                Main_Screen.toggle_changer_available  = True
                Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                IO.output(record.green1_led, False)
                IO.output(record.green2_led, False)
                IO.output(record.green3_led, False)
                IO.output(record.green4_led, False)
                IO.output(record.green5_led, False)
                IO.output(record.yellow_led, False)
                IO.output(record.red1_led, False)
                IO.output(record.red2_led, False)
                IO.output(record.red3_led, False)
                IO.output(record.red4_led, False)
                IO.output(record.red5_led, False)
                IO.output(record.gps_working, False)
                sense.end1StLeg = 0 # Clear the value of the D2 cell on the spreadsheet 
                record.numLeg = 0    # Clear the value of the C2 cell on the spreadsheet
                time.sleep(1)
                main_loop()
                input_loop()
                static_message_loop()
                race_setup_screen_loop()
                #Main_Screen.screen1()
                #Main_Screen.screen2()
                #Main_Screen.screen3()
                #Main_Screen.screen4()
        t = threading.Thread(target = save_excel)
        t.start()    

class Save_File_Text_Input_Screen(Screen):
    save_text_file = ObjectProperty()
    text = StringProperty('')
    file_to_save = ""

    def get_name_of_the_file(self):
        
        self.text = self.save_text_file.text
        reg_file = re.split(r'^\s+', self.save_text_file.text)
        if len(reg_file) == 1:
            if len(reg_file[0]) == 0:
                print("NoTHING")
                Save_File_Text_Input_Screen.file_to_save = ""
                guiApp.message_selected = Static_Message.INVALID_NAME
            else:
                #print("take this name as name of the file")
                Save_File_Text_Input_Screen.file_to_save = reg_file[0]
                print("I will load this file: {}".format(Save_File_Text_Input_Screen.file_to_save))
        else:
            for i in reg_file:
                if len(i) > 0:
                    Save_File_Text_Input_Screen.file_to_save = i
                    print("the name of the file is: {}".format(i))
                    print("I will load this file: {}".format(Save_File_Text_Input_Screen.file_to_save))

class Race_Setup_Screen(Screen):
    update_val = Show_Val()
    startRecordPulse = False
    time_to_clear = 0
    clearStartButton = False

    timePush = 0
    timeC = 0
    timeDiff = 0
    keyToMeasure = False

    def update(self, dt):
        
        
        #if Race_Setup_Screen.clearStartButton == True:
        #    currentTime = int(time.time() - Race_Setup_Screen.time_to_clear)
        #    if currentTime >= 3:
        #        print("current time {}".format(currentTime))
        #        print("clear the start button")
        if record.disablePulseMeas == True:
            self.update_val.hidePulseMeasure = True
        else:
            self.update_val.hidePulseMeasure = False
        if Race_Setup_Screen.startRecordPulse == True:
            if record.clear_timer == 1:
                self.update_val.pulseCounter = str(sense.pulses) + " Pulses"
            #Race_Setup_Screen.keyToMeasure = True
            Race_Setup_Screen.timeC = time.time() 
            Race_Setup_Screen.timePush = int(Race_Setup_Screen.timeC - Race_Setup_Screen.timeDiff)
            #print("time pushed {}".format(Race_Setup_Screen.timePush))
            if Race_Setup_Screen.clearStartButton == True:
                if Race_Setup_Screen.timePush >= 3:
                    self.update_val.pulseCounter = "START RECORD"
                    Race_Setup_Screen.clearStartButton = False
                    Race_Setup_Screen.startRecordPulse = False
                    sense.startReadPulses = False
                            #record.toggle_status = 0
                    sense.pulses = 0
                    record.record = False
                    
                    del routes.waypoint_latitude[:]
                    del routes.waypoint_longitude[:]
                    record.tm = 0
                    record.record = False
                    record.recording_status_flag = False
                    record.toggle_status_flag = False
                    record.t = 0
                    record.recorded_time["seconds"] = 0.000
                    record.recorded_time["minutes"] = 0
                    record.recorded_time["hours"] = 0
                    record.clear_timer = 0
                    record.set_record_time = 0.001
                    record.toggle_button_time = 0.0
                    del routes.waypoint_hours[:]
                    del routes.waypoint_minutes[:]
                    del routes.waypoint_seconds[:]
                    record.flag_go = 0
                    #sense.pulses = 0
                    sense.mph = 0
                    Main_Screen.spIndex = 0
                    sense.factor = 1
                    sense.decimal_factor = 1
                    sense.lastPulse = 0
                    Main_Screen.lock_toggle = False
                    sense.last_ = 0
                    Main_Screen.stable_gps = time.time()
                    sense.race_started = False
                    routes.counter = 0.0
                    sense.dist_meas = 0.0
                    routes.last_dist_meas = 0.0
                    record.general_purpose = 0
                    del routes.waypoint_mph[:]
                    del routes.waypoint_dist_meas[:]
                    del routes.waypoint_counter[:]
                    del routes.waypoint_comparison_data[:]
                    Main_Screen.update_val.current_value = "0"
                    Main_Screen.update_val.current_color = Main_Screen.yellow
                    Main_Screen.update_val.dist_travel_2 =  "0" + "mi"
                    Main_Screen.update_val.dist_travel = "0" + " mi"
                    #Main_Screen.update_val.count_down_up = "0" + " MPH"
                    Main_Screen.update_val.count_down_up = "0" + " Secs"
                    Main_Screen.update_val.timer_update = "0:0:0.0"
                    Main_Screen.flag_race_mode = 0
                    routes.last_dist_meas = 0
                    Main_Screen.my_iterator = 0
                    Routes_List_Screen.preloaded_route_done = 1 
                    Routes_List_Screen.uploaded = 0
                    Main_Screen.result = 0
                    record.starter_switch = 0
                    Main_Screen.toggle_changer_available  = True
                    Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                    IO.output(record.green1_led, False)
                    IO.output(record.green2_led, False)
                    IO.output(record.green3_led, False)
                    IO.output(record.green4_led, False)
                    IO.output(record.green5_led, False)
                    IO.output(record.yellow_led, False)
                    IO.output(record.red1_led, False)
                    IO.output(record.red2_led, False)
                    IO.output(record.red3_led, False)
                    IO.output(record.red4_led, False)
                    IO.output(record.red5_led, False)
                    IO.output(record.gps_working, False)
        else:
            self.update_val.pulseCounter = "START RECORD"
                #self.update_val.pulseCounter = "START RECORD"
        if Race_Setup_Screen.keyToMeasure == False:
            Race_Setup_Screen.timeDiff = time.time() 
        
        
        #print("print from race setup screen")
        #global STOP_BUTTON
        #if STOP_BUTTON == 1:
        #    self.update_val.speed_target_indicator = str(Input_Target_Speed_Screen.t_speed_variable)
    
    def stop_record(self):
        print("stop record")
        if Race_Setup_Screen.startRecordPulse == True:
            Race_Setup_Screen.clearStartButton = True
        #Race_Setup_Screen.startRecordPulse = False
        
        record.clear_timer = 0

        
        #p = pulseFunc.CountLogger()
        #p = 

        #p = pulseFunc.CountLogger(sense.pulses)
        #print("this is star record and value is p:  {}".format(p.myPulse))
        #global STOP_BUTTON
        #STOP_BUTTON = 0

    def start_record(self):
        #Race_Setup_Screen.startRecordPulse = True
        if Race_Setup_Screen.startRecordPulse == False:
            Race_Setup_Screen.startRecordPulse = True
            sense.startReadPulses = True
        if Race_Setup_Screen.clearStartButton == False:
            record.clear_timer = 0
            if record.clear_timer == 0:
                if record.toggle_status == 1:
                    record.seconds_from_gps = gps.seconds
                    record.toggle_status = 0
                    record.clear_timer = 1
                    record.starter_switch = 1
                else:
                    record.seconds_from_gps = gps.seconds
                    record.toggle_status = 1
                    record.clear_timer = 1
                    record.starter_switch = 2
        Race_Setup_Screen.keyToMeasure = True
        #record.toggle_status = 0
        if Race_Setup_Screen.clearStartButton == True:
            Race_Setup_Screen.time_to_clear = time.time()

        
    def detach(self):
        Race_Setup_Screen.keyToMeasure = False
        
class Static_Message(Screen):
    update_val = Show_Val()
    INVALID_NAME                    = 0
    SAVED_IN_SYSTEM_AND_USB         = 1
    SAVED_IN_SYSTEM                 = 2
    NO_SELECTED_ROUTE               = 3
    NO_FILE_SELECTED                = 4
    ROUTE_LOADED                    = 5
    INVALID_FORMAT                  = 6
    DELETED_ROUTE                   = 7
    NO_ROUTE_SELECTED_IN_SAVE_MODE  = 8
    ROUTE_SAVED_IN_USB              = 9
    NO_USB_FOUND                    = 10
    DELETED_ROUTE_USB               = 11
    TIRE_DATA_SAVED                 = 12
    DELETED_TIRE_DATA               = 13
    INCORRECT_SPEED_FORMAT          = 14
    INCORRECT_INCH_FORMAT           = 15
    INCORRECT_TEMP_FORMAT           = 16
    INCORRECT_PRESS_FORMAT          = 17
    SAVING_LABEL                    = 18
    LOADING_LABEL                   = 19
    NO_ERROR_IN_USB                 = 254
    NO_ERROR                        = 255

    def update(self, dt):
        if guiApp.message_selected == Static_Message.SAVED_IN_SYSTEM_AND_USB:
            self.update_val.message = "ROUTE SUCCESFULLY SAVED \nIN SYSTEM AND USB DRIVE"
        elif guiApp.message_selected == Static_Message.SAVED_IN_SYSTEM:
            self.update_val.message = "ROUTE SUCCESFULLY \n   SAVED IN SYSTEM"
        elif guiApp.message_selected == Static_Message.INVALID_NAME:
            self.update_val.message = "ERROR. PLEASE, INPUT A \nVALID NAME TO THE FILE"
        elif guiApp.message_selected == Static_Message.NO_FILE_SELECTED:
            self.update_val.message = "ERROR. PLEASE, SELECT\n A ROUTE"
        elif guiApp.message_selected == Static_Message.SAVING_LABEL:
            self.update_val.message = "SAVING FILE, PLEASE WAIT..."
        elif guiApp.message_selected == Static_Message.LOADING_LABEL:
            self.update_val.message = "LOADING FILE, PLEASE WAIT..."
        elif guiApp.message_selected == Static_Message.ROUTE_LOADED:
            self.update_val.message = "ROUTE SUCCESSFULLY \n               LOADED"
        elif guiApp.message_selected == Static_Message.INVALID_FORMAT:
            self.update_val.message = "ERROR. INVALID ROUTE FORMAT"
        elif guiApp.message_selected == Static_Message.DELETED_ROUTE:
            self.update_val.message = "ROUTE SUCCESFULLY \n             DELETED"   
        elif guiApp.message_selected == Static_Message.NO_ROUTE_SELECTED_IN_SAVE_MODE:
            self.update_val.message = "ERROR. PLEASE, \nSELECT A ROUTE"   
        elif guiApp.message_selected == Static_Message.ROUTE_SAVED_IN_USB:
            self.update_val.message = "ROUTE SUCCESFULLY \n       SAVED IN USB"   
        elif guiApp.message_selected == Static_Message.NO_USB_FOUND:
            self.update_val.message = "ERROR. USB DRIVE \n       NOT FOUND"
        elif guiApp.message_selected == Static_Message.DELETED_ROUTE_USB:
            self.update_val.message = "ROUTE SUCCESFULLY \n             DELETED"  
        elif guiApp.message_selected == Static_Message.TIRE_DATA_SAVED:
            self.update_val.message = "TIRE DATA IS SUCCESFULLY \n          SAVED IN SYSTEM"
        elif guiApp.message_selected == Static_Message.DELETED_TIRE_DATA:
            self.update_val.message = "ROUTE SUCCESFULLY \n   SAVED IN SYSTEM"
        elif guiApp.message_selected == Static_Message.INCORRECT_SPEED_FORMAT:
            self.update_val.message = "ERROR. PLEASE, INPUT\n A CORRECT VALUE"
        elif guiApp.message_selected == Static_Message.INCORRECT_INCH_FORMAT:
            self.update_val.message = "ERROR. PLEASE, INPUT\n A CORRECT VALUE"
        elif guiApp.message_selected == Static_Message.INCORRECT_TEMP_FORMAT:
            self.update_val.message = "ERROR. PLEASE, INPUT\n A CORRECT VALUE"
        elif guiApp.message_selected == Static_Message.INCORRECT_PRESS_FORMAT:
            self.update_val.message = "ERROR. PLEASE, INPUT\n A CORRECT VALUE"
            
class Confirmation_Screen_To_Load(Screen):  # screen to confirm to load a route while you are in route list screen
    preloaded_target_speed = 0
    #waypoint_mph_from_datalogger =       []
    waypoint_dist_meas_from_datalogger = []
    #waypoint_latitude_from_datalogger =  []
    #waypoint_longitude_from_datalogger = []
    waypoint_time                      = []
    #waypoint_hours_from_datalogger =     []
    #waypoint_minutes_from_datalogger =   []
    #waypoint_seconds_from_datalogger =   []
    #waypoint_counter_from_datalogger =   []
    waypoint_target_speed_from_datalogger = []
    old_target_speed = 0
    counter_from_datalogger_from_datalogger = 0
    last_dist_meas_from_datalogger = 0.0

    preloaded_route_done = 1
    uploaded = 0

    my_time_array = "0"
    my_time_complete = []

    excel_load_indicator = 0
    list_2 = []
    file_loaded = ""

    show_message = False

    file_to_delete = ""
    file_to_load = ""
    file_already_loaded = False
    student_list = ObjectProperty()
    key_file_list = False

    tire_size = 0
    
    @classmethod
    def set_tire_rotation(cls, tire_size):
        cls.tire_size = tire_size

    @classmethod
    def get_tire_rotation(cls):
        return cls.tire_size 

    def load_file(self):
        
        Save_File_Text_Input_Screen.message_manager = Static_Message.LOADING_LABEL
        guiApp.message_selected = Static_Message.LOADING_LABEL
        def load_route():
            print("load the file: {}".format(Routes_List_Screen.file_to_load))
            try:
            # If a list item is selected
                if Routes_List_Screen.file_to_load:
                    my_path = "/home/pi/Documents/LoadRoute/"
                    if len(Routes_List_Screen.file_to_load) == 0:
                        pass
                    else:
                        data_name = Routes_List_Screen.file_to_load + ".xlsx"
                        Routes_List_Screen.file_loaded = Routes_List_Screen.file_to_load

                        wb = openpyxl.load_workbook(my_path + data_name, read_only = True, data_only = True)
                        ws = wb.active

                        del Routes_List_Screen.waypoint_dist_meas_from_datalogger[:]
                        del Routes_List_Screen.waypoint_time[:]
                        sense.end1StLeg = 0  # Clear the value of the D2 cell on the spreadsheet 
                        record.numLeg = 0    # Clear the value of the C2 cell on the spreadsheet

                        sense.end1StLeg = ws['E2'] # Read the value of the D2 cell on the spreadsheet 
                        sense.end1StLeg = sense.end1StLeg.value

                        record.numLeg = ws['D2']    # Read the value of the C2 cell on the spreadsheet
                        record.numLeg = record.numLeg.value

                        # coords for point A
                        # sense.coords_a_x = float(ws['D4'].value)
                        # sense.coords_a_y = float(ws['D5'].value)
                        # sense.coords_b_x = float(ws['E4'].value)
                        # sense.coords_b_y = float(ws['E5'].value)
                        # print("A: {}, {}, type:{}, {}".format(sense.coords_a_x, sense.coords_a_y, type(sense.coords_a_x), type(sense.coords_a_y)))
                        # print("B: {}, {}, type:{}, {}".format(sense.coords_b_x, sense.coords_b_y, type(sense.coords_b_x), type(sense.coords_b_y)))

                        print("the number of leg is: {}".format(record.numLeg))
                        print("type of the leg is: {}".format(type(record.numLeg)))



                        # Read each cells from A column, we need the lenght to measure how much reads we need to do
                        for row in ws.iter_rows('A{}:A{}'.format(2, ws.max_row)):  # Set min and max 
                            for cell in row:
                                if cell.value != None:
                                    try:
                                        Routes_List_Screen.waypoint_dist_meas_from_datalogger.append(float(cell.value))
                                    except Exception as e:
                                        print("error here: {}".format(e))

                        if not Routes_List_Screen.waypoint_dist_meas_from_datalogger:
                            guiApp.message_selected = Static_Message.INVALID_FORMAT
                        
                        else:
                            self.set_tire_rotation(Routes_List_Screen.waypoint_dist_meas_from_datalogger[0])
                            sense.decimal_factor = self.get_tire_rotation()
                            print("tire rotation is: {}".format(self.get_tire_rotation()))
                            # Read each cells from B column, we need the lenght to measure how much reads we need to do
                            for row in ws.iter_rows('B{}:B{}'.format(3, (len(Routes_List_Screen.waypoint_dist_meas_from_datalogger) + 2))):
                                for cell in row:
                                    #if cell == isinstance(record.numLeg, unicode) == True and 
                                    #print("data: {} type of data: {}".format(cell.value, type(cell.value)))
                                    Routes_List_Screen.waypoint_time.append(cell.value)
                            print("loaded pulses {}".format(len(Routes_List_Screen.waypoint_dist_meas_from_datalogger)))
                            print("time loaded {}".format(len(Routes_List_Screen.waypoint_time)))
                            # exception handler for when the number leg is not valid 
                            if record.numLeg == 1 and isinstance(record.numLeg, unicode) == False or record.numLeg == "" or record.numLeg == " " or record.numLeg == None:
                                print("the number of leg is: {} and type {}".format(record.numLeg, type(record.numLeg)))
                                Routes_List_Screen.preloaded_route_done = 0 # set to 0 a preloaded route
                                Routes_List_Screen.uploaded = 1
                                self.change_name(Routes_List_Screen.file_to_load)
                                Routes_List_Screen.file_already_loaded = True
                                guiApp.message_selected = Static_Message.ROUTE_LOADED
                            elif record.numLeg == 2 and record.numLeg != None and isinstance(record.numLeg, unicode) == False:
                                print("it's 2nd leg")
                                if sense.end1StLeg != None and isinstance(sense.end1StLeg, unicode) == False:
                                    print("end of leg is: {} and type {}".format(sense.end1StLeg, type(sense.end1StLeg)))
                                    Routes_List_Screen.preloaded_route_done = 0 # set to 0 a preloaded route
                                    Routes_List_Screen.uploaded = 1
                                    self.change_name(Routes_List_Screen.file_to_load)
                                    Routes_List_Screen.file_already_loaded = True
                                    guiApp.message_selected = Static_Message.ROUTE_LOADED
                                else:
                                    sense.end1StLeg = 0
                                    record.numLeg = 0
                                    if Routes_List_Screen.file_already_loaded == True:
                                        pass
                                    else:
                                        Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                                    guiApp.message_selected = Static_Message.INVALID_FORMAT
                            else:
                                sense.end1StLeg = 0
                                record.numLeg = 0
                                if Routes_List_Screen.file_already_loaded == True:
                                    pass
                                else:
                                    Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                                guiApp.message_selected = Static_Message.INVALID_FORMAT

            except KeyError:
                sense.end1StLeg = 0
                record.numLeg = 0
                if Routes_List_Screen.file_already_loaded == True:
                    pass
                else:
                    Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                guiApp.message_selected = Static_Message.INVALID_FORMAT

            except ValueError:
                sense.end1StLeg = 0
                record.numLeg = 0
                if Routes_List_Screen.file_already_loaded == True:
                    pass
                else:
                    Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                guiApp.message_selected = Static_Message.INVALID_FORMAT

            except TypeError:
                sense.end1StLeg = 0
                record.numLeg = 0
                if Routes_List_Screen.file_already_loaded == True:
                    pass
                else:
                    Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
                guiApp.message_selected = Static_Message.INVALID_FORMAT
        
        t = threading.Thread(target = load_route)
        t.start()   
        #except IOError:
        #    sense.end1StLeg = 0
        #    record.numLeg = 0
        #    if Routes_List_Screen.file_already_loaded == True:
        #        pass
        #    else:
        #        Main_Screen.update_val.text_gps = "NO ROUTE \nSELECTED"
        #    guiApp.message_selected = Static_Message.INVALID_FORMAT

    def change_name(self, name_to_show):
        print("{}".format(len(name_to_show)))
        if name_to_show: # if file is selected
            if len(name_to_show) == 0:
                pass
            else:
                name = name_to_show
                if len(name) >= 9:
                    print("little name {}...".format(name[:7]))
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name[:7] + "..."
                else:
                    Main_Screen.update_val.text_gps = "    ROUTE\nSELECTED:\n" + name
                print("name {}".format(name))

class Confirmation_Screen_To_Delete(Screen):    # screen to confirm to delete a route in load route list screen
    
    def delete_spreadsheet(self):
        print("file i want to delete: {}".format(Routes_List_Screen.file_to_delete))
        # If a list item is selected
        if Routes_List_Screen.file_to_delete:
            # Get the text from the item selected
            selection = Routes_List_Screen.file_to_delete
            deleted_file = "rm /home/pi/Documents/LoadRoute/" + '"' + selection + '"' + ".xlsx"
            # Remove the matching item
            print("selection {}".format(selection))
            subprocess.call(deleted_file, shell = True)
            guiApp.message_selected = Static_Message.DELETED_ROUTE
        else:
            guiApp.message_selected = Static_Message.NO_SELECTED_ROUTE

class Confirmation_Save_List_Screen(Screen):  # screen to confirm to save a route in usb drive menu
    
    def save_file(self):
        global STOP_BUTTON
        key = False
        out = subprocess.check_output("df", shell = True)   # Save devices plugged to device into "out"
        if(len(out) >= 570):                                # Check if the string is enoug to know if a drive is connected 
            print("looking for a usb drive")
            splited = out.split(" ")                        # Split the string every space
            for i in range(0, len(splited)):                # iterate the list
                if(len(splited[i]) == 24):                  # if the item of the list is same as target item "sda1" device
                    usb_dev = splited[i].split("/")         # split the string every "/"
                    for i in range(0, len(usb_dev)):        # iterate current list
                        if((usb_dev[i] == "sda1") or (usb_dev[i] == "sdb1")):           # if iterated item is sda1 or sdb1
                            print("{}".format(usb_dev[i]))
                            check_files = subprocess.check_output("ls /media/pi", shell = True) # save the output of 'ls /media/pi'
                            check_files = check_files.split("\n")                               # removing the "\n"
                            folder = subprocess.check_output("ls /media/pi/" + '"' + check_files[0] + '"', shell = True) # save the output of 'ls /media/pi'
                            folder = folder.split("\n")   
                            for i in range(len(folder)):
                                print(folder[i])
                                if folder[i] == "GPS SYSTEM GENERATED ROUTE":
                                    print("I found the file {}".format(folder[i]))
                                    key = True
                                    break

                            if key == False:
                                folder_create = 'mkdir /media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'"
                                print("folder doesn't exist, I will create it")
                                print("{}".format(folder_create))
                                subprocess.call(folder_create, shell = True)
                                route = "/media/pi/" + '"' + check_files[0] + '"' + "/"
                                path = "cp /home/pi/Documents/SaveRoute/" + "'" + File_List_To_Save_Usb.file_name + '.xlsx' + "'"
                                destiny_folder = '/media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'" + "/"
                                print("folder to save 1: {}".format(path + " " + destiny_folder))
                                subprocess.call(path + " " + destiny_folder, shell = True)
                                guiApp.message_selected = Static_Message.ROUTE_SAVED_IN_USB
                            else:
                                key = False
                                path = "cp /home/pi/Documents/SaveRoute/" + "'" + File_List_To_Save_Usb.file_name + '.xlsx' + "'"
                                destiny_folder = '/media/pi/' + '"' + check_files[0] + '"' + "/" + "'" + "GPS SYSTEM GENERATED ROUTE" + "'" + "/"
                                print("folder to save 2: {}".format(path + " " + destiny_folder))
                                subprocess.call(path + " " + destiny_folder, shell = True)
                                guiApp.message_selected = Static_Message.ROUTE_SAVED_IN_USB
                        else:
                            guiApp.message_selected = Static_Message.NO_USB_FOUND
        else:
            guiApp.message_selected = Static_Message.NO_USB_FOUND

class Confirmation_Delete_List_Screen(Screen):    # screen to confirm to delete a route in load route list screen
    
    def delete_spreadsheet(self):
        # If a list item is selected
        if File_List_To_Save_Usb.file_name:
            # Get the text from the item selected
            selection = File_List_To_Save_Usb.file_name
            deleted_file = "rm /home/pi/Documents/SaveRoute/" + '"' + selection + '"' + ".xlsx"
            # Remove the matching item
            print("selection {}".format(selection))
            subprocess.call(deleted_file, shell = True)
            guiApp.message_selected = Static_Message.DELETED_ROUTE_USB
        else:
            guiApp.message_selected = Static_Message.NO_SELECTED_ROUTE

class Screen_Management(ScreenManager):
    mc = ObjectProperty(None)
    loading_screen = ObjectProperty(None)

presentation = 0 # Builder.load_file("gui.kv")

class guiApp(App):
    message_selected = 0

    def build(self):
        global main_loop, input_loop, static_message_loop, race_setup_screen_loop
        screen_1 = Main_Screen()
        screen_3 = Static_Message()
        screen_4 = Race_Setup_Screen()
        main_loop = Clock.schedule_interval(screen_1.update, 1.0/10.0)
        static_message_loop = Clock.schedule_interval(screen_3.update, 1.0)
        race_setup_screen_loop = Clock.schedule_interval(screen_4.update, 1.0)
        sense.gps_thread = threading.Thread(target=gps_thread)  # instance the thread
        sense.gps_thread.start()   # call to start the thread
        self.m = Screen_Management(transition = FadeTransition()) # 
        return presentation

    def show_main_screen(self, dt):
        self.m.current = "main_screen_name"
        
def gps_thread():
    try:
        global thread_flag
        global STOP_BUTTON
        #global clock_status 
       
        while thread_flag == False:
            #clock_var_status = len(str(clock_status.get_callback()))
            #if clock_var_status < 53:
            #    print("clock is dead")
            #print(" clock status special {}".format(clock_var_status))
            #print("hello")
            #clock_status.cancel()
            #clock_status()
            #Main_Screen.screen_1()
            gps.update_gnss_data(init_gps)  # get gps data
            #print("decimal_factor {} and factor {}".format(sense.decimal_factor, sense.factor))
            #if Race_Setup_Screen.startRecordPulse == True and TIRE_RADIUS_REF != 0 and sense.decimal_factor != 0:
            #    record.record = True
            if STOP_BUTTON == 1:
                Main_Screen.stable_gps_result = int(time.time() - Main_Screen.stable_gps)
            if record.recording_status_flag == True:
                if sense.switch_mph_reference == False:   
                    if STOP_BUTTON == 1:
                        if sense.elapse != 0: # This condition is to prevent zero division
                            if TIRE_RADIUS_REF != 0:
                                if sense.decimal_factor != 0:
                                      # to know which type of race will be (1 or 2 leg)
                                    if record.numLeg == 1 or record.numLeg == None:
                                        if sense.pulses >= round(sense.decimal_factor, 0):
                                            # print("current pulses: {} target pulses {} and distance {}")
                                            difference = sense.pulses - round(sense.decimal_factor, 0)
                                            if difference > 0:
                                                sense.pulses = sense.pulses - difference
                                            elif difference < 0:
                                                sense.pulses = sense.pulses + (-difference)
                                            #print("dist meas!!")
                                            sense.dist_meas += 0.1
                                            sense.pulses += 1 
                                            sense.factor += 1
                                            # sense.decimal_factor = round((sense.tire_rotations_per_mile * sense.factor), 0)
                                            sense.decimal_factor = (Confirmation_Screen_To_Load.get_tire_rotation() * sense.factor)
                                            print("decimal factor: {}".format(sense.decimal_factor))
                                        sense.dist_meas = round(sense.dist_meas, 2) # Rounded 2 decimals
                                        if(Routes_List_Screen.preloaded_route_done == 0 and Routes_List_Screen.uploaded == 1): 
                                            if Main_Screen.spIndex < len(Routes_List_Screen.waypoint_dist_meas_from_datalogger):
                                                if (sense.pulses >= Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex]) and record.flag_go == 1:
                                                    if (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == None) or (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == " "):
                                                        pass
                                                    else:
                                                        Main_Screen.comparison = record.get_sec(init_record, str(Routes_List_Screen.waypoint_time[Main_Screen.spIndex])) - round(Main_Screen.current_race_time, 2)  
                                                        #print("comparison!!! {}".format(Main_Screen.comparison))
                                                        #print("current time: {}".format(Main_Screen.current_race_time))
                                                    #print("recorded time {}:{}:{}".format(int(record.recorded_time["hours"]),
                                                    #    int(record.recorded_time["minutes"]),
                                                    #    round(record.recorded_time["seconds"], 3)))
                                                    Main_Screen.final_time = round(Main_Screen.current_race_time, 2)
                                                    Main_Screen.record_time = str(datetime.timedelta(seconds = Main_Screen.final_time))
                                                    Main_Screen.record_time = Main_Screen.record_time[:10]
                                                    #print("other recorded time to log {}".format(Main_Screen.record_time))
                                                    #print("time {}".format(Main_Screen.final_time))
                                                    routes.get_data(
                                                        routes_data,
                                                        sense.pulses,#round(record.mph_accurate, 3),
                                                        int(record.mph_accurate),
                                                        round(sense.dist_meas, 1),
                                                        gps.latitude,
                                                        gps.longitude,
                                                        Main_Screen.record_time)
                                                    Main_Screen.flag_race_mode = 1
                                                    sense.aux_dist += 0.2
                                                    Main_Screen.spIndex += 1

                                    elif record.numLeg == 2 or record.numLeg != 0:
                                        if sense.pulses >= round(sense.decimal_factor, 0):
                                            print("start measure ")
                                            difference = sense.pulses - round(sense.decimal_factor, 0)
                                            if difference > 0:
                                                sense.pulses = sense.pulses - difference
                                            elif difference < 0:
                                                sense.pulses = sense.pulses + (-difference)
                                            sense.dist_meas += 0.1
                                            sense.pulses += 1 
                                            sense.factor += 1
                                            # sense.decimal_factor = round((sense.tire_rotations_per_mile * sense.factor), 0)
                                            sense.decimal_factor = (Confirmation_Screen_To_Load.get_tire_rotation() * sense.factor)
                                            print("decimal factor: {}".format(sense.decimal_factor))
                                        sense.dist_meas = round(sense.dist_meas, 2) # Rounded 2 decimals
                                        if(Routes_List_Screen.preloaded_route_done == 0 and Routes_List_Screen.uploaded == 1): 
                                            # print("start measure 22")
                                            if Main_Screen.spIndex < len(Routes_List_Screen.waypoint_dist_meas_from_datalogger):
                                                # print("start measure 333")
                                                
                                                # print("waypoint: {} - record.flag_go {}".format(Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex], record.flag_go))
                                                if (sense.pulses >= Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex]) and record.flag_go == 1:
                                                    
                                                    print("start measure 4")
                                                    print("pulses {} - waypoint {} - decimal factor: {}".format(sense.pulses, Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex], sense.decimal_factor))
                                                    if (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == None) or (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == " "):
                                                        pass
                                                    else:
                                                        print("start measure 5")
                                                        Main_Screen.comparison = record.get_sec(init_record, str(Routes_List_Screen.waypoint_time[Main_Screen.spIndex])) - round(Main_Screen.current_race_time, 2)  

                                                    Main_Screen.final_time = round(Main_Screen.current_race_time, 2)
                                                    Main_Screen.record_time = str(datetime.timedelta(seconds = Main_Screen.final_time))
                                                    Main_Screen.record_time = Main_Screen.record_time[:10]
                                                    routes.get_data(
                                                        routes_data,
                                                        sense.pulses,#round(record.mph_accurate, 3),
                                                        #int(record.mph_accurate),
                                                        round(sense.dist_meas, 1),
                                                        gps.latitude,
                                                        gps.longitude,
                                                        Main_Screen.record_time)
                                                    Main_Screen.flag_race_mode = 1
                                                    sense.aux_dist += 0.2
                                                    Main_Screen.spIndex += 1 
                                                    Main_Screen.key_1st_leg_end = False

                                                if sense.pulses > sense.end1StLeg and Main_Screen.bool_leg_already_elapsed == False and Main_Screen.key_1st_leg_end == True:
                                                    STOP_BUTTON = 0
                                                    sense.pulses = sense.end1StLeg
                                                    Main_Screen.end1stLegTime = time.time()
                                                    Main_Screen.bool_race_leg = True
                                                    Main_Screen.bool_leg_already_elapsed = True
                                                    Main_Screen.add_2nd_leg = True
                                                    Main_Screen.key_1st_leg_end = False
                else:
                    if STOP_BUTTON == 1:
                        if sense.elapse != 0: # This condition is to prevent zero division
                            if TIRE_RADIUS_REF != 0:
                                # PLACE TO TEST THE GPS functionality during race time 
                                print("A: {}, {}, type:{}, {}".format(sense.coords_a_x, sense.coords_a_y, type(sense.coords_a_x), type(sense.coords_a_y)))
                                print("B: {}, {}, type:{}, {}".format(sense.coords_b_x, sense.coords_b_y, type(sense.coords_b_x), type(sense.coords_b_y)))

                                if sense.decimal_factor != 0:
                                      # to know which type of race will be (1 or 2 leg)
                                    if record.numLeg == 1 or record.numLeg == None:
                                        if sense.pulses >= round(sense.decimal_factor, 0):
                                            #print("current pulses: {} target pulses {} and distance {}")
                                            difference = sense.pulses - round(sense.decimal_factor, 0)
                                            if difference > 0:
                                                sense.pulses = sense.pulses - difference
                                            elif difference < 0:
                                                sense.pulses = sense.pulses + (-difference)
                                            #print("dist meas!!")
                                            sense.dist_meas += 0.1
                                            sense.pulses += 1 
                                            sense.factor += 1
                                            sense.decimal_factor = round((sense.tire_rotations_per_mile * sense.factor), 0)
                                        sense.dist_meas = round(sense.dist_meas, 2) # Rounded 2 decimals
                                        if(Routes_List_Screen.preloaded_route_done == 0 and Routes_List_Screen.uploaded == 1): 
                                            if Main_Screen.spIndex < len(Routes_List_Screen.waypoint_dist_meas_from_datalogger):
                                                if (sense.pulses >= Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex]) and record.flag_go == 1:
                                                    if (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == None) or (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == " "):
                                                        pass
                                                    else:
                                                        Main_Screen.comparison = record.get_sec(init_record, str(Routes_List_Screen.waypoint_time[Main_Screen.spIndex])) - round(Main_Screen.current_race_time, 2)  
                                                        #print("comparison!!! {}".format(Main_Screen.comparison))
                                                        #print("current time: {}".format(Main_Screen.current_race_time))
                                                    #print("recorded time {}:{}:{}".format(int(record.recorded_time["hours"]),
                                                    #    int(record.recorded_time["minutes"]),
                                                    #    round(record.recorded_time["seconds"], 3)))
                                                    Main_Screen.final_time = round(Main_Screen.current_race_time, 2)
                                                    Main_Screen.record_time = str(datetime.timedelta(seconds = Main_Screen.final_time))
                                                    Main_Screen.record_time = Main_Screen.record_time[:10]
                                                    #print("other recorded time to log {}".format(Main_Screen.record_time))
                                                    #print("time {}".format(Main_Screen.final_time))
                                                    routes.get_data(
                                                        routes_data,
                                                        sense.pulses,#round(record.mph_accurate, 3),
                                                        int(record.mph_accurate),
                                                        round(sense.dist_meas, 1),
                                                        gps.latitude,
                                                        gps.longitude,
                                                        Main_Screen.record_time)
                                                    Main_Screen.flag_race_mode = 1
                                                    sense.aux_dist += 0.2
                                                    Main_Screen.spIndex += 1

                                    elif record.numLeg == 2 or record.numLeg != 0:
                                        if sense.pulses >= round(sense.decimal_factor, 0):

                                            difference = sense.pulses - round(sense.decimal_factor, 0)
                                            if difference > 0:
                                                sense.pulses = sense.pulses - difference
                                            elif difference < 0:
                                                sense.pulses = sense.pulses + (-difference)
                                            sense.dist_meas += 0.1
                                            sense.pulses += 1 
                                            sense.factor += 1
                                            sense.decimal_factor = round((sense.tire_rotations_per_mile * sense.factor), 0)
                                        sense.dist_meas = round(sense.dist_meas, 2) # Rounded 2 decimals
                                        if(Routes_List_Screen.preloaded_route_done == 0 and Routes_List_Screen.uploaded == 1): 
                                            if Main_Screen.spIndex < len(Routes_List_Screen.waypoint_dist_meas_from_datalogger):
                                                
                                                if (sense.pulses >= Routes_List_Screen.waypoint_dist_meas_from_datalogger[Main_Screen.spIndex]) and record.flag_go == 1:
                                                    
                                                    if (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == None) or (Routes_List_Screen.waypoint_time[Main_Screen.spIndex] == " "):
                                                        pass
                                                    else:
                                                        Main_Screen.comparison = record.get_sec(init_record, str(Routes_List_Screen.waypoint_time[Main_Screen.spIndex])) - round(Main_Screen.current_race_time, 2)  

                                                    Main_Screen.final_time = round(Main_Screen.current_race_time, 2)
                                                    Main_Screen.record_time = str(datetime.timedelta(seconds = Main_Screen.final_time))
                                                    Main_Screen.record_time = Main_Screen.record_time[:10]
                                                    routes.get_data(
                                                        routes_data,
                                                        sense.pulses,#round(record.mph_accurate, 3),
                                                        #int(record.mph_accurate),
                                                        round(sense.dist_meas, 1),
                                                        gps.latitude,
                                                        gps.longitude,
                                                        Main_Screen.record_time)
                                                    Main_Screen.flag_race_mode = 1
                                                    sense.aux_dist += 0.2
                                                    Main_Screen.spIndex += 1 
                                                    Main_Screen.key_1st_leg_end = False

                                                if sense.pulses > sense.end1StLeg and Main_Screen.bool_leg_already_elapsed == False and Main_Screen.key_1st_leg_end == True:
                                                    STOP_BUTTON = 0
                                                    sense.pulses = sense.end1StLeg
                                                    Main_Screen.end1stLegTime = time.time()
                                                    Main_Screen.bool_race_leg = True
                                                    Main_Screen.bool_leg_already_elapsed = True
                                                    Main_Screen.add_2nd_leg = True
                                                    Main_Screen.key_1st_leg_end = False

    except KeyboardInterrupt:
        thread_flag = True
        IO.cleanup()
        sys.exit
    
    except IOError:
        pass

try:
    guiApp().run()
except IOError:
    print "IO Error"